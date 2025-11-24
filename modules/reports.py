#models/reports.py
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import mm, cm
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
from datetime import datetime
from typing import Dict, List, Optional  # ✅ AGREGAR Optional aquí
import os
import time
import sys
import subprocess
import platform

if platform.system() == "Windows":
    try:
        import win32print
        import win32api
    except ImportError:
        win32print = None
        win32api = None
        print("Advertencia: pywin32 no está instalado. La impresión en Windows puede fallar. Instale con: pip install pywin32")

from typing import Dict, List
from modules.models import obtener_recibo_por_id, obtener_configuracion, obtener_recibos_por_folio

# ✅ AGREGAR ESTOS IMPORTS PARA LA FUNCIÓN DE EXCEL
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
# ==================== CONFIGURACIÓN DE RECIBO ====================

# IMPORTANTE: Recibo en formato 1/3 carta - ORIENTACIÓN VERTICAL
RECIBO_ANCHO = 21.6 * cm
RECIBO_ALTO = 9.1 * cm

# Ruta del logo
LOGO_PATH = os.path.join('assets', 'lagoo.png')

# ==================== UTILIDADES DE IMPRESIÓN (Windows) ====================

def _buscar_sumatra() -> str | None:
    """
    Busca SumatraPDF en rutas comunes (x64/x86) y retorna la ruta si existe.
    """
    posibles = [
        r"C:\Program Files\SumatraPDF\SumatraPDF.exe",
        r"C:\Program Files (x86)\SumatraPDF\SumatraPDF.exe",
        os.path.expandvars(r"%LOCALAPPDATA%\SumatraPDF\SumatraPDF.exe"),
        os.path.expandvars(r"%LOCALAPPDATA%\Microsoft\WindowsApps\SumatraPDF.exe"),
    ]
    for p in posibles:
        if os.path.exists(p):
            return p
    return None

def _imprimir_pdf_windows(ruta_pdf: str, impresora: str | None = None) -> None:
    """
    Imprime en Windows con múltiples fallbacks:
    1) SumatraPDF (mejor opción)
    2) Abrir PDF y dejar que el usuario imprima manualmente
    """
    import subprocess
    
    # 1) Intentar con SumatraPDF
    sumatra = _buscar_sumatra()
    if sumatra:
        try:
            args = [sumatra]
            if impresora:
                args += ["-print-to", impresora]
            else:
                args += ["-print-to-default"]
            args += ["-exit-on-print", ruta_pdf]
            subprocess.run(args, check=True, timeout=10)
            print(f"✓ Impreso con SumatraPDF: {ruta_pdf}")
            return
        except Exception as e:
            print(f"⚠ SumatraPDF falló: {e}")
    
    # 2) Fallback: Solo ABRIR el PDF (el usuario imprime manualmente)
    try:
        os.startfile(ruta_pdf)
        print(f"⚠ PDF abierto para impresión manual: {ruta_pdf}")
        # Mostrar mensaje al usuario
        import tkinter as tk
        from tkinter import messagebox
        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo(
            "Impresión Manual", 
            "El PDF se ha abierto.\n\n"
            "Por favor, presiona Ctrl+P para imprimir manualmente.\n\n"
            "Recomendación: Instala SumatraPDF para impresión automática."
        )
        root.destroy()
        return
    except Exception as e:
        raise RuntimeError(
            f"No se puede imprimir en Windows.\n\n"
            f"Soluciones:\n"
            f"1. Instala SumatraPDF (https://www.sumatrapdfreader.org/)\n"
            f"2. Configura una impresora predeterminada en Windows\n"
            f"3. Asocia PDFs con Adobe Reader\n\n"
            f"Error: {e}"
        )


# ==================== GENERACIÓN DE RECIBOS ====================

def generar_recibo_pdf(recibo_id: int, es_reimpresion: bool = False) -> str:
    """Genera el PDF de un recibo en formato 1/3 carta - VERTICAL"""
    recibo = obtener_recibo_por_id(recibo_id)
    if not recibo:
        raise ValueError("Recibo no encontrado")

    nombre_oficina = obtener_configuracion('nombre_oficina') or 'ASOCIACIÓN DE RIEGO'
    ubicacion = obtener_configuracion('ubicacion') or 'Tezontepec de Aldama, Hgo.'

    recibos_dir = os.path.join('database', 'recibos')
    os.makedirs(recibos_dir, exist_ok=True)

    sufijo = '_REIMPRESION' if es_reimpresion else ''
    filename = f"recibo_{recibo['folio']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{sufijo}.pdf"
    filepath = os.path.join(recibos_dir, filename)

    # Crear canvas con tamaño en orientación VERTICAL
    c = canvas.Canvas(filepath, pagesize=(RECIBO_ANCHO, RECIBO_ALTO))
    _dibujar_recibo_principal(c, recibo, nombre_oficina, ubicacion, es_reimpresion)
    c.save()

    return filepath

def generar_recibo_pdf_temporal(recibo_id: int, es_reimpresion: bool = False) -> str:
    """
    Genera el PDF de un recibo TEMPORAL que será eliminado después de imprimir.
    Agrupa recibos con el mismo folio para mostrar todos los riegos.
    """
    # 1. Obtener datos del recibo inicial
    recibo = obtener_recibo_por_id(recibo_id)
    if not recibo:
        raise ValueError("Recibo no encontrado")

    # 2. Buscar si hay más recibos con el mismo folio
    recibos_folio = obtener_recibos_por_folio(recibo['folio'])
    
    # 3. Datos base (del primer recibo)
    # Usamos copy para no modificar el original si fuera necesario, aunque aquí no importa mucho
    datos_base = recibos_folio[0].copy()
    
    # 4. Agregar números de riego y sumar costos
    numeros_riego = sorted([r['numero_riego'] for r in recibos_folio])
    # Formato: "1, 2, 3"
    texto_riegos = ", ".join(map(str, numeros_riego))
    
    costo_total = sum(r['costo'] for r in recibos_folio)
    
    # 5. Actualizar datos para el reporte
    datos_base['numero_riego'] = texto_riegos
    datos_base['costo'] = costo_total

    nombre_oficina = obtener_configuracion('nombre_oficina') or 'ASOCIACIÓN DE RIEGO'
    ubicacion = obtener_configuracion('ubicacion') or 'Tezontepec de Aldama, Hgo.'

    # Crear carpeta temporal en /tmp (Mac/Linux) o %TEMP% (Windows)
    if platform.system() == "Windows":
        temp_dir = os.path.join(os.environ.get('TEMP', os.getcwd()), 'recibos_temp')
    else:
        temp_dir = os.path.join('/tmp', 'recibos_temp')

    os.makedirs(temp_dir, exist_ok=True)

    sufijo = '_REIMPRESION' if es_reimpresion else ''
    filename = f"recibo_{datos_base['folio']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{sufijo}.pdf"
    filepath = os.path.join(temp_dir, filename)

    c = canvas.Canvas(filepath, pagesize=(RECIBO_ANCHO, RECIBO_ALTO))
    _dibujar_recibo_principal(c, datos_base, nombre_oficina, ubicacion, es_reimpresion)
    c.save()

    return filepath

# ==================== IMPRESIÓN (TEMPORALES) ====================

def imprimir_recibo_y_limpiar(pdf_path: str):
    """
    Intenta imprimir un archivo PDF en Windows o macOS y luego lo elimina.
    """
    sistema = platform.system()
    print(f"Intentando imprimir en {sistema} desde: {pdf_path}")

    if not os.path.exists(pdf_path):
        print(f"Error: El archivo PDF no existe: {pdf_path}")
        return

    impreso = False

    if sistema == "Windows":
        if win32print and win32api:
            try:
                # Usar win32api para imprimir
                win32api.ShellExecute(0, "print", pdf_path, None, ".", 0)
                print(f"Comando de impresión enviado para win32: {pdf_path}")
                impreso = True
            except Exception as e:
                 print(f"Error con win32api.ShellExecute: {e}")
        else:
             print("pywin32 no disponible. Intentando alternativas...")

        if not impreso:
            # Alternativa menos confiable: os.startfile
            try:
                os.startfile(pdf_path, "print") # Esta acción puede no ser silenciosa
                print(f"Comando de impresión enviado para os.startfile: {pdf_path}")
                impreso = True
            except Exception as e:
                 print(f"Error con os.startfile: {e}")

    elif sistema == "Darwin": # macOS
        try:
            # Usar 'lp' para imprimir en macOS (requiere CUPS instalado, que generalmente lo está)
            subprocess.run(["lp", pdf_path], check=True)
            print(f"PDF impreso en macOS usando lp: {pdf_path}")
            impreso = True
        except subprocess.CalledProcessError as e:
            print(f"Error al imprimir con 'lp' en macOS: {e}")
        except FileNotFoundError:
            print("Comando 'lp' no encontrado en macOS. Verifique la instalación de CUPS o la ruta.")

    else: # Linux u Otro
        try:
            # Usar 'lp' para imprimir en Linux (requiere CUPS instalado)
            subprocess.run(["lp", pdf_path], check=True)
            print(f"PDF impreso en Linux usando lp: {pdf_path}")
            impreso = True
        except subprocess.CalledProcessError as e:
            print(f"Error al imprimir con 'lp' en Linux: {e}")
        except FileNotFoundError:
            print("Comando 'lp' no encontrado. Verifique la instalación de CUPS o la ruta.")

    # Esperar un momento para que el comando de impresión se procese
    import time
    time.sleep(1) # Ajusta si es necesario

    # Eliminar el archivo temporal después del intento de impresión
    try:
        os.remove(pdf_path)
        print(f"Archivo temporal eliminado: {pdf_path}")
    except OSError as e:
        print(f"Error al eliminar archivo temporal {pdf_path}: {e}")

# ==================== DIBUJO DE RECIBO ====================

def _dibujar_recibo_principal(c, recibo: Dict, nombre_oficina: str, ubicacion: str, es_reimpresion: bool):
    """
    Dibuja el recibo principal - MARCA DE AGUA AL FINAL
    """
    
    # ===== COLORES =====
    COLOR_VERDE = colors.HexColor('#B8D1BF')
    COLOR_BEIGE = colors.HexColor('#FFFFFF')
    COLOR_BEIGE_OSCURO = colors.HexColor('#C9B99A')
    COLOR_TEXTO = colors.HexColor('#2C3E2E')
    COLOR_TEXTO_GRIS = colors.HexColor('#666666')
    
    # ===== FONDO BEIGE =====
    c.setFillColor(COLOR_BEIGE)
    c.roundRect(0.15*cm, 0.15*cm, RECIBO_ANCHO - 0.3*cm, RECIBO_ALTO - 0.3*cm, 
                0.5*cm, stroke=0, fill=1)
    
    # ===== HEADER VERDE (más alto) =====
    c.setFillColor(COLOR_VERDE)
    c.roundRect(0.4*cm, RECIBO_ALTO - 2.3*cm, RECIBO_ANCHO - 0.8*cm, 1.9*cm, 
                0.4*cm, stroke=0, fill=1)
    
    # ===== LOGO (más grande) =====
    if os.path.exists(LOGO_PATH):
        try:
            c.drawImage(LOGO_PATH, 0.7*cm, RECIBO_ALTO - 2.1*cm, 
                       width=1.7*cm, height=1.7*cm, mask='auto')
        except:
            pass
    
    # ===== TÍTULO (texto más grande) =====
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 10)
    c.drawCentredString(RECIBO_ANCHO/2 + 0.5*cm, RECIBO_ALTO - 1*cm, 
                       "ASOCIACIÓN DE CAMPESINOS DE BOMBEO Y REBOMBEO")
    
    c.setFont("Helvetica-Bold", 9)
    c.drawCentredString(RECIBO_ANCHO/2 + 0.5*cm, RECIBO_ALTO - 1.35*cm, 
                       "DEL CERRO DEL XICUCO A.C. M7-1")
    
    c.setFont("Helvetica", 8)
    c.drawCentredString(RECIBO_ANCHO/2 + 0.5*cm, RECIBO_ALTO - 1.75*cm, 
                       "RFC: ACB030619G68")
    
    # ===== SEPARADOR =====
    y_pos = RECIBO_ALTO - 2.45*cm
    c.setStrokeColor(COLOR_BEIGE_OSCURO)
    c.setLineWidth(0.5)
    c.line(0.7*cm, y_pos, RECIBO_ANCHO - 0.7*cm, y_pos)
    
    # ===== CAJA DE DATOS (más alta) =====
    y_pos -= 0.2*cm
    c.setFillColor(colors.white)
    c.roundRect(0.7*cm, y_pos - 1.5*cm, RECIBO_ANCHO - 1.4*cm, 1.4*cm, 
                0.25*cm, stroke=1, fill=1)
    
    # ===== GRID DE DATOS (texto más grande) =====
    c.setFillColor(COLOR_TEXTO)
    c.setFont("Helvetica-Bold", 8.5)
    
    col1 = 1*cm
    col2 = 6.5*cm
    col3 = 12*cm
    col4 = 17.5*cm
    
    row_y = y_pos - 0.45*cm
    
    # FILA 1
    c.drawString(col1, row_y, f"NO. RECIBO: {recibo['folio']}")
    c.drawString(col2, row_y, f"No. Lote: {recibo['numero_lote']}")
    c.drawString(col3, row_y, f"No. Riego: {recibo['numero_riego']}")
    c.drawString(col4, row_y, f"Barrio: {recibo['barrio']}")
    
    row_y -= 0.45*cm
    
    # FILA 2
    col1_fila2 = 1*cm
    col2_fila2 = 8.5*cm
    col3_fila2 = 16*cm
    
    c.drawString(col1_fila2, row_y, f"Cultivo: {recibo['cultivo']}")
    c.drawString(col2_fila2, row_y, f"Superficie: {recibo['superficie']} ha")
    c.drawString(col3_fila2, row_y, f"Ciclo: {recibo['ciclo']}")
    
    # ===== RECIBÍ DE (más espacio y texto más grande) =====
    y_pos = row_y - 0.95*cm
    c.setFillColor(COLOR_TEXTO_GRIS)
    c.setFont("Helvetica", 8.5)
    c.drawString(0.8*cm, y_pos, "Recibí de:")
    
    c.setFillColor(COLOR_TEXTO)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2.2*cm, y_pos, recibo['nombre'].upper())
    
    # ===== CONCEPTO (más espacio) =====
    y_pos -= 0.5*cm
    c.setStrokeColor(COLOR_BEIGE_OSCURO)
    c.line(0.8*cm, y_pos, RECIBO_ANCHO - 0.8*cm, y_pos)
    
    y_pos -= 0.3*cm
    c.setFillColor(COLOR_TEXTO_GRIS)
    c.setFont("Helvetica", 8)
    c.drawString(0.8*cm, y_pos, "Concepto: Pago de cuota de riego para el ciclo agrícola")
    
    # ===== TOTAL + MONTO (más espacio y texto más grande) =====
    y_pos -= 0.5*cm
    
    c.setFillColor(COLOR_TEXTO)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(0.8*cm, y_pos, "TOTAL")
    
    # Caja del monto (más grande)
    monto_x = RECIBO_ANCHO - 5*cm
    c.setFillColor(colors.white)
    c.setStrokeColor(COLOR_VERDE)
    c.setLineWidth(1.5)
    c.roundRect(monto_x, y_pos - 0.35*cm, 4.2*cm, 0.75*cm, 
                0.25*cm, stroke=1, fill=1)
    
    c.setFillColor(COLOR_TEXTO_GRIS)
    c.setFont("Helvetica", 6.5)
    c.drawString(monto_x + 0.2*cm, y_pos + 0.15*cm, "(pago en efectivo)")
    
    c.setFillColor(COLOR_TEXTO)
    c.setFont("Helvetica-Bold", 15)
    c.drawRightString(monto_x + 4*cm, y_pos - 0.15*cm, f"${recibo['costo']:.2f}")
    
    # ===== FOOTER (más espacio) =====
    y_pos -= 0.9*cm
    c.setStrokeColor(COLOR_BEIGE_OSCURO)
    c.setLineWidth(0.5)
    c.line(0.8*cm, y_pos, RECIBO_ANCHO - 0.8*cm, y_pos)
    
    y_pos -= 0.3*cm
    c.setFillColor(COLOR_TEXTO_GRIS)
    c.setFont("Helvetica", 7.5)
    
    fecha_obj = datetime.strptime(recibo['fecha'], '%Y-%m-%d')
    c.drawString(0.8*cm, y_pos, 
                f"C. Juan Aldama #25, Col. Centro, Tezontepec de Aldama. Fecha: {fecha_obj.strftime('%d/%m/%Y')}")
    
    y_pos -= 0.28*cm
    hora_obj = datetime.strptime(recibo['hora'], '%H:%M:%S')
    am_pm = "p.m." if hora_obj.hour >= 12 else "a.m."
    hora_12 = hora_obj.hour if hora_obj.hour <= 12 else hora_obj.hour - 12
    if hora_12 == 0:
        hora_12 = 12
    
    c.drawString(0.8*cm, y_pos, f"Hora: {hora_12:02d}:{hora_obj.minute:02d}:{hora_obj.second:02d} {am_pm}")
    
    # Firma (más espacio)
    c.drawRightString(RECIBO_ANCHO - 0.8*cm, y_pos + 0.28*cm, "Firma Recaudador")
    c.line(RECIBO_ANCHO - 4*cm, y_pos + 0.18*cm, RECIBO_ANCHO - 0.8*cm, y_pos + 0.18*cm)
    
    # ===== LEYENDA LEGAL (texto más grande y más espacio) =====
    y_pos -= 0.45*cm
    c.setFont("Helvetica", 6)
    
    c.drawString(0.7*cm, y_pos, 
                "Este recibo ampara el pago de cuota ordinaria destinada exclusivamente al mantenimiento y operación del módulo de riego, conforme al régimen fiscal")
    y_pos -= 0.22*cm
    c.drawString(0.7*cm, y_pos,
                "de personas morales con fines no lucrativos. Exento de IVA y de ISR conforme a los artículos 79 y 80 de la Ley del ISR y al artículo 15, fracción XII de la Ley del IVA.")
    
    # ===== MARCA DE AGUA (CENTRADA VERTICAL Y HORIZONTALMENTE) =====
    if es_reimpresion:
        c.saveState()
        c.setFont("Helvetica-Bold", 32)
        c.setFillColorRGB(0.9, 0.9, 0.9)
        c.rotate(30)
        c.drawString(8 * cm, 0.1 * cm, "REIMPRESIÓN") # Cambiado de 0.5*cm a 0.1*cm
        c.restoreState()

# ==================== REPORTE DIARIO ====================

def generar_reporte_diario(fecha: str, recibos: List[Dict]) -> str:
    """Genera un reporte PDF del día con todos los recibos - ORIENTACIÓN VERTICAL"""
    reportes_dir = os.path.join('database', 'reportes')
    os.makedirs(reportes_dir, exist_ok=True)
    filename = f"reporte_diario_{fecha.replace('-', '')}.pdf"
    filepath = os.path.join(reportes_dir, filename)

    # Canvas con tamaño PORTRAIT (vertical) - letter = (8.5 x 11 inches)
    c = canvas.Canvas(filepath, pagesize=letter)

    nombre_oficina = obtener_configuracion('nombre_oficina') or 'ASOCIACIÓN DE RIEGO'

    # --- Añadir Logo al Reporte Diario ---
    if os.path.exists(LOGO_PATH):
        try:
            logo_width = 2 * cm
            logo_height = 2 * cm
            c.drawImage(LOGO_PATH, 2*cm, letter[1] - 3*cm, width=logo_width, height=logo_height, mask='auto')
        except Exception as e:
            print(f"Error al añadir logo al reporte diario: {e}")
    # --------------------------------------

    y_pos = letter[1] - 2*cm
    margen_izq = 2*cm
    margen_der = letter[0] - 2*cm

    # ENCABEZADO
    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(letter[0]/2, y_pos, nombre_oficina.upper())
    y_pos -= 0.7*cm

    c.setFont("Helvetica-Bold", 12)
    fecha_obj = datetime.strptime(fecha, '%Y-%m-%d')
    c.drawCentredString(letter[0]/2, y_pos, f"REPORTE DIARIO - {fecha_obj.strftime('%d/%m/%Y')}")
    y_pos -= 0.7*cm

    c.setFont("Helvetica", 10)
    c.drawCentredString(letter[0]/2, y_pos, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    y_pos -= 1*cm

    # TABLA DE RECIBOS
    if not recibos:
        c.setFont("Helvetica", 11)
        c.drawCentredString(letter[0]/2, y_pos, "No hay recibos registrados en este día")
    else:
        c.setFont("Helvetica-Bold", 8)
        col_widths = [1.5*cm, 2*cm, 1.5*cm, 5*cm, 2.5*cm, 1.5*cm, 2.5*cm, 2*cm]
        col_x = [margen_izq]
        for w in col_widths[:-1]:
            col_x.append(col_x[-1] + w)

        headers = ["Folio", "Hora", "Lote", "Nombre", "Cultivo", "Riego", "Acción", "Monto"]
        for i, header in enumerate(headers):
            c.drawString(col_x[i], y_pos, header)
        y_pos -= 0.3*cm
        c.line(margen_izq, y_pos, margen_der, y_pos)
        y_pos -= 0.4*cm

        c.setFont("Helvetica", 7)
        total_dia = 0

        for recibo in recibos:
            if y_pos < 3*cm:
                c.showPage()
                # Volver a dibujar el logo en la nueva página si es necesario
                if os.path.exists(LOGO_PATH):
                    try:
                        logo_width = 2 * cm
                        logo_height = 2 * cm
                        c.drawImage(LOGO_PATH, 2*cm, letter[1] - 3*cm, width=logo_width, height=logo_height, mask='auto')
                    except Exception as e:
                        print(f"Error al añadir logo a nueva página del reporte: {e}")

                y_pos = letter[1] - 2*cm
                c.setFont("Helvetica-Bold", 8)
                for i, header in enumerate(headers):
                    c.drawString(col_x[i], y_pos, header)
                y_pos -= 0.3*cm
                c.line(margen_izq, y_pos, margen_der, y_pos)
                y_pos -= 0.4*cm
                c.setFont("Helvetica", 7)

            c.drawString(col_x[0], y_pos, str(recibo['folio']))
            c.drawString(col_x[1], y_pos, recibo['hora'][:5])
            c.drawString(col_x[2], y_pos, recibo['numero_lote'])
            nombre = recibo['nombre'][:30] if len(recibo['nombre']) > 30 else recibo['nombre']
            c.drawString(col_x[3], y_pos, nombre)
            c.drawString(col_x[4], y_pos, recibo['cultivo'])
            c.drawString(col_x[5], y_pos, str(recibo['numero_riego']))
            tipo = "Nueva" if recibo['tipo_accion'] == 'Nueva siembra' else "Adicional"
            c.drawString(col_x[6], y_pos, tipo)
            c.drawRightString(col_x[7] + 2*cm, y_pos, f"${recibo['costo']:.2f}")
            total_dia += recibo['costo']
            y_pos -= 0.35*cm

        y_pos -= 0.2*cm
        c.line(margen_izq, y_pos, margen_der, y_pos)
        y_pos -= 0.5*cm

        c.setFont("Helvetica-Bold", 11)
        c.drawString(margen_izq, y_pos, f"TOTAL DEL DÍA:")
        c.drawRightString(margen_der, y_pos, f"${total_dia:.2f}")
        y_pos -= 0.5*cm

        c.setFont("Helvetica", 9)
        c.drawString(margen_izq, y_pos, f"Total de recibos emitidos: {len(recibos)}")

    c.save()
    return filepath

def generar_reporte_mensual_pdf(anio: int, mes: int, recibos: List[Dict]) -> str:
    """Genera un reporte PDF mensual con todos los recibos - ORIENTACIÓN VERTICAL"""
    reportes_dir = os.path.join('database', 'reportes')
    os.makedirs(reportes_dir, exist_ok=True)
    filename = f"reporte_mensual_{anio}_{mes:02d}.pdf"
    filepath = os.path.join(reportes_dir, filename)

    # Canvas con tamaño PORTRAIT (vertical)
    c = canvas.Canvas(filepath, pagesize=letter)

    nombre_oficina = obtener_configuracion('nombre_oficina') or 'ASOCIACIÓN DE RIEGO'

    # --- Añadir Logo ---
    if os.path.exists(LOGO_PATH):
        try:
            logo_width = 2 * cm
            logo_height = 2 * cm
            c.drawImage(LOGO_PATH, 2*cm, letter[1] - 3*cm, width=logo_width, height=logo_height, mask='auto')
        except Exception as e:
            print(f"Error al añadir logo al reporte mensual: {e}")
    # -------------------

    y_pos = letter[1] - 2*cm
    margen_izq = 2*cm
    margen_der = letter[0] - 2*cm

    # ENCABEZADO
    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(letter[0]/2, y_pos, nombre_oficina.upper())
    y_pos -= 0.7*cm

    c.setFont("Helvetica-Bold", 12)
    meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
             "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    nombre_mes = meses[mes-1]
    c.drawCentredString(letter[0]/2, y_pos, f"REPORTE MENSUAL - {nombre_mes.upper()} {anio}")
    y_pos -= 0.7*cm

    c.setFont("Helvetica", 10)
    c.drawCentredString(letter[0]/2, y_pos, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    y_pos -= 1*cm

    # TABLA DE RECIBOS
    if not recibos:
        c.setFont("Helvetica", 11)
        c.drawCentredString(letter[0]/2, y_pos, "No hay recibos registrados en este mes")
    else:
        c.setFont("Helvetica-Bold", 8)
        col_widths = [1.5*cm, 2*cm, 1.5*cm, 5*cm, 2.5*cm, 1.5*cm, 2.5*cm, 2*cm]
        col_x = [margen_izq]
        for w in col_widths[:-1]:
            col_x.append(col_x[-1] + w)

        headers = ["Folio", "Fecha", "Lote", "Nombre", "Cultivo", "Riego", "Acción", "Monto"]
        for i, header in enumerate(headers):
            c.drawString(col_x[i], y_pos, header)
        y_pos -= 0.3*cm
        c.line(margen_izq, y_pos, margen_der, y_pos)
        y_pos -= 0.4*cm

        c.setFont("Helvetica", 7)
        total_mes = 0

        for recibo in recibos:
            if y_pos < 3*cm:
                c.showPage()
                if os.path.exists(LOGO_PATH):
                    try:
                        c.drawImage(LOGO_PATH, 2*cm, letter[1] - 3*cm, width=2*cm, height=2*cm, mask='auto')
                    except: pass

                y_pos = letter[1] - 2*cm
                c.setFont("Helvetica-Bold", 8)
                for i, header in enumerate(headers):
                    c.drawString(col_x[i], y_pos, header)
                y_pos -= 0.3*cm
                c.line(margen_izq, y_pos, margen_der, y_pos)
                y_pos -= 0.4*cm
                c.setFont("Helvetica", 7)

            c.drawString(col_x[0], y_pos, str(recibo['folio']))
            c.drawString(col_x[1], y_pos, recibo['fecha']) # Mostrar fecha en lugar de hora
            c.drawString(col_x[2], y_pos, recibo['numero_lote'])
            nombre = recibo['nombre'][:30] if len(recibo['nombre']) > 30 else recibo['nombre']
            c.drawString(col_x[3], y_pos, nombre)
            c.drawString(col_x[4], y_pos, recibo['cultivo'])
            c.drawString(col_x[5], y_pos, str(recibo['numero_riego']))
            tipo = "Nueva" if recibo['tipo_accion'] == 'Nueva siembra' else "Adicional"
            c.drawString(col_x[6], y_pos, tipo)
            c.drawRightString(col_x[7] + 2*cm, y_pos, f"${recibo['costo']:.2f}")
            total_mes += recibo['costo']
            y_pos -= 0.35*cm

        y_pos -= 0.2*cm
        c.line(margen_izq, y_pos, margen_der, y_pos)
        y_pos -= 0.5*cm

        c.setFont("Helvetica-Bold", 11)
        c.drawString(margen_izq, y_pos, f"TOTAL DEL MES:")
        c.drawRightString(margen_der, y_pos, f"${total_mes:.2f}")
        y_pos -= 0.5*cm

        c.setFont("Helvetica", 9)
        c.drawString(margen_izq, y_pos, f"Total de recibos emitidos: {len(recibos)}")

    c.save()
    return filepath

# ==================== IMPRESIÓN DIRECTA (NO TEMPORALES) ====================

def imprimir_recibo(ruta_pdf: str, impresora: str = None):
    """Envía el PDF a la impresora sin eliminar el archivo (para PDFs no temporales)"""
    if not os.path.exists(ruta_pdf):
        raise FileNotFoundError(f"Archivo no encontrado: {ruta_pdf}")

    sistema = platform.system()

    try:
        if sistema == 'Windows':
            _imprimir_pdf_windows(ruta_pdf, impresora)
        elif sistema == 'Darwin':
            cmd = ['lp', ruta_pdf] if not impresora else ['lp', '-d', impresora, ruta_pdf]
            subprocess.run(cmd, check=True)
        else:
            cmd = ['lp', ruta_pdf] if not impresora else ['lp', '-d', impresora, ruta_pdf]
            subprocess.run(cmd, check=True)
        return True
    except Exception as e:
        print(f"Error al imprimir: {e}")
        return False

# ==================== LISTA DE IMPRESORAS ====================

def obtener_impresoras_disponibles() -> List[str]:
    """
    Lista impresoras disponibles:
    - Windows: pywin32 si está, si no PowerShell/WMIC (fallback).
    - macOS/Linux: lpstat.
    """
    try:
        sistema = platform.system()

        if sistema == "Windows":
            # 1) Intento con pywin32
            try:
                import win32print  # type: ignore
                impresoras = win32print.EnumPrinters(2)
                return [imp[2] for imp in impresoras if len(imp) >= 3]
            except Exception:
                pass  # seguir al fallback

            # 2) PowerShell (Get-CimInstance)
            try:
                ps = [
                    "powershell", "-NoProfile", "-Command",
                    "Get-CimInstance -ClassName Win32_Printer | Select-Object -ExpandProperty Name"
                ]
                r = subprocess.run(ps, capture_output=True, text=True, timeout=5)
                names = [ln.strip() for ln in r.stdout.splitlines() if ln.strip()]
                if names:
                    return names
            except Exception:
                pass

            # 3) WMIC (legacy pero aún frecuente)
            try:
                r = subprocess.run(["wmic", "printer", "get", "name"],
                                   capture_output=True, text=True, timeout=5)
                names = [ln.strip() for ln in r.stdout.splitlines()[1:] if ln.strip()]
                if names:
                    return names
            except Exception:
                pass

            return ["Impresora por defecto"]

        elif sistema == "Darwin":  # macOS
            r = subprocess.run(["lpstat", "-p", "-d"], capture_output=True, text=True, timeout=5)
            impresoras = []
            for ln in r.stdout.splitlines():
                if ln.startswith("printer"):
                    partes = ln.split()
                    if len(partes) >= 2:
                        impresoras.append(partes[1])
            return impresoras if impresoras else ["Impresora por defecto"]

        else:  # Linux
            r = subprocess.run(["lpstat", "-p", "-d"], capture_output=True, text=True, timeout=5)
            impresoras = []
            for ln in r.stdout.splitlines():
                if ln.startswith("printer"):
                    partes = ln.split()
                    if len(partes) >= 2:
                        impresoras.append(partes[1])
            return impresoras if impresoras else ["Impresora por defecto"]

    except Exception:
        return ["Impresora por defecto"]

# ==================== ABRIR PDF ====================

def abrir_pdf(ruta_pdf: str):
    """Abre el PDF con el visor predeterminado del sistema"""
    if not os.path.exists(ruta_pdf):
        raise FileNotFoundError(f"Archivo no encontrado: {ruta_pdf}")

    sistema = platform.system()

    try:
        if sistema == 'Windows':
            os.startfile(ruta_pdf)  # depende de asociación, solo para ver
        elif sistema == 'Darwin':
            subprocess.run(['open', ruta_pdf])
        else:
            subprocess.run(['xdg-open', ruta_pdf])
        return True
    except Exception as e:
        print(f"Error al abrir PDF: {e}")
        return False

# ==================== EXPORTACIÓN A EXCEL ====================

def exportar_a_excel(recibos: List[Dict], filename: str) -> str:
    """Exporta una lista de recibos a un archivo Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        # Crear libro de trabajo
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Recibos"

        # Encabezados
        headers = ['Folio', 'Fecha', 'Hora', 'Lote', 'Nombre', 'Localidad', 'Barrio',
                   'Superficie', 'Cultivo', 'Riego No.', 'Acción', 'Costo']

        # Estilo de encabezados
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Datos
        for row, recibo in enumerate(recibos, 2):
            ws.cell(row=row, column=1, value=recibo['folio'])
            ws.cell(row=row, column=2, value=recibo['fecha'])
            ws.cell(row=row, column=3, value=recibo['hora'])
            ws.cell(row=row, column=4, value=recibo['numero_lote'])
            ws.cell(row=row, column=5, value=recibo['nombre'])
            ws.cell(row=row, column=6, value=recibo['localidad'])
            ws.cell(row=row, column=7, value=recibo['barrio'])
            ws.cell(row=row, column=8, value=recibo['superficie'])
            ws.cell(row=row, column=9, value=recibo['cultivo'])
            ws.cell(row=row, column=10, value=recibo['numero_riego'])
            ws.cell(row=row, column=11, value=recibo['tipo_accion'])
            ws.cell(row=row, column=12, value=recibo['costo'])

        # Ajustar anchos de columna
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            adjusted_width = min((max_length + 2), 50)
            ws.column_dimensions[column].width = adjusted_width

        # Guardar archivo
        reportes_dir = os.path.join('database', 'reportes')
        os.makedirs(reportes_dir, exist_ok=True)
        filepath = os.path.join(reportes_dir, filename)
        wb.save(filepath)

        return filepath

    except ImportError:
        raise ImportError("La librería 'openpyxl' no está instalada. Instálala con: pip install openpyxl")
    except Exception as e:
        raise Exception(f"Error al exportar a Excel: {e}")
    
def generar_corte_caja_excel(fecha: str, recibos: List[Dict]) -> str:
    """
    Genera un archivo Excel con el corte de caja del día.
    
    Columnas en orden:
    FOLIO | NO DE LOTE | CICLO | BARRIO | CULTIVO | SUPERFICIE DE LA MILPA | 
    RIEGO DE LA MILPA | CAMPESINO | CUANTO SE COBRA | FECHA
    
    Args:
        fecha: Fecha en formato YYYY-MM-DD
        recibos: Lista de recibos del día
        
    Returns:
        Ruta del archivo Excel generado
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Crear directorio si no existe
    reportes_dir = os.path.join('database', 'reportes')
    os.makedirs(reportes_dir, exist_ok=True)
    
    # Nombre del archivo
    fecha_obj = datetime.strptime(fecha, '%Y-%m-%d')
    fecha_str = fecha_obj.strftime('%Y%m%d')
    nombre_archivo = f"corte_caja_{fecha_str}.xlsx"
    ruta_excel = os.path.join(reportes_dir, nombre_archivo)
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Corte de Caja"
    
    # ===== ESTILOS =====
    titulo_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    titulo_fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
    
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    total_font = Font(name='Calibri', size=12, bold=True)
    total_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== ENCABEZADO =====
    ws.merge_cells('A1:J1')
    cell_titulo = ws['A1']
    cell_titulo.value = 'CORTE DE CAJA'
    cell_titulo.font = titulo_font
    cell_titulo.fill = titulo_fill
    cell_titulo.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:J2')
    cell_fecha = ws['A2']
    cell_fecha.value = f"Fecha: {fecha_obj.strftime('%d/%m/%Y')}"
    cell_fecha.font = Font(name='Calibri', size=12, bold=True)
    cell_fecha.alignment = Alignment(horizontal='center')
    
    nombre_oficina = obtener_configuracion('nombre_oficina') or 'SISTEMA DE RIEGO'
    ws.merge_cells('A3:J3')
    cell_oficina = ws['A3']
    cell_oficina.value = nombre_oficina
    cell_oficina.font = Font(name='Calibri', size=11, italic=True)
    cell_oficina.alignment = Alignment(horizontal='center')
    
    # ===== CABECERAS (10 COLUMNAS) =====
    headers = ['FOLIO', 'NO. LOTE', 'CICLO', 'BARRIO', 'CULTIVO', 'SUP', 
               'RIEGO', 'RECIBI DE', 'SERVICIO', 'FECHA']
    
    row_num = 5
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # ===== DATOS (10 COLUMNAS) =====
    row_num = 6
    total_monto = 0
    
    for recibo in recibos:
        if recibo.get('eliminado'):
            continue
        
        # Columna 1: FOLIO
        ws.cell(row=row_num, column=1, value=recibo['folio']).border = border
        
        # Columna 2: NO DE LOTE
        ws.cell(row=row_num, column=2, value=recibo['numero_lote']).border = border
        
        # Columna 3: CICLO
        ws.cell(row=row_num, column=3, value=recibo['ciclo']).border = border
        
        # Columna 4: BARRIO
        ws.cell(row=row_num, column=4, value=recibo['barrio']).border = border
        
        # Columna 5: CULTIVO
        ws.cell(row=row_num, column=5, value=recibo['cultivo']).border = border
        
        # Columna 6: SUPERFICIE DE LA MILPA
        sup_cell = ws.cell(row=row_num, column=6, value=recibo['superficie'])
        sup_cell.border = border
        sup_cell.alignment = Alignment(horizontal='right')
        
        # Columna 7: RIEGO DE LA MILPA
        ws.cell(row=row_num, column=7, value=recibo['numero_riego']).border = border
        
        # Columna 8: CAMPESINO
        ws.cell(row=row_num, column=8, value=recibo['nombre']).border = border
        
        # Columna 9: CUANTO SE COBRA
        monto_cell = ws.cell(row=row_num, column=9, value=recibo['costo'])
        monto_cell.border = border
        monto_cell.number_format = '$#,##0.00'
        monto_cell.alignment = Alignment(horizontal='right')
        
        # Columna 10: FECHA
        ws.cell(row=row_num, column=10, value=recibo['fecha']).border = border
        
        total_monto += recibo['costo']
        row_num += 1
    
    # ===== TOTALES =====
    row_num += 1
    ws.merge_cells(f'A{row_num}:H{row_num}')
    cell_total_label = ws.cell(row=row_num, column=1)
    cell_total_label.value = 'TOTAL DEL DÍA:'
    cell_total_label.font = total_font
    cell_total_label.fill = total_fill
    cell_total_label.alignment = Alignment(horizontal='right')
    cell_total_label.border = border
    
    cell_total_monto.fill = total_fill
    cell_total_monto.number_format = '$#,##0.00'
    cell_total_monto.alignment = Alignment(horizontal='right')
    cell_total_monto.border = border
    
    # Ajustar anchos
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        
    wb.save(ruta_excel)
    return ruta_excel

def generar_reporte_mensual_excel(anio: int, mes: int, recibos: List[Dict]) -> str:
    """
    Genera un archivo Excel con el reporte mensual.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    reportes_dir = os.path.join('database', 'reportes')
    os.makedirs(reportes_dir, exist_ok=True)
    
    filename = f"reporte_mensual_{anio}_{mes:02d}.xlsx"
    filepath = os.path.join(reportes_dir, filename)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Reporte {mes:02d}-{anio}"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Título
    meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
             "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    ws['A1'] = f"REPORTE MENSUAL DE VENTAS - {meses[mes-1].upper()} {anio}"
    ws['A1'].font = Font(size=14, bold=True)
    
    # Encabezados
    headers = ['FOLIO', 'FECHA', 'HORA', 'LOTE', 'NOMBRE', 'CULTIVO', 'RIEGO', 'ACCIÓN', 'COSTO']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    total = 0
    for row, recibo in enumerate(recibos, 4):
        ws.cell(row=row, column=1, value=recibo['folio'])
        ws.cell(row=row, column=2, value=recibo['fecha'])
        ws.cell(row=row, column=3, value=recibo['hora'])
        ws.cell(row=row, column=4, value=recibo['numero_lote'])
        ws.cell(row=row, column=5, value=recibo['nombre'])
        ws.cell(row=row, column=6, value=recibo['cultivo'])
        ws.cell(row=row, column=7, value=recibo['numero_riego'])
        ws.cell(row=row, column=8, value=recibo['tipo_accion'])
        ws.cell(row=row, column=9, value=recibo['costo']).number_format = '$#,##0.00'
        total += recibo['costo']
        
    # Total
    row_total = len(recibos) + 5
    ws.cell(row=row_total, column=8, value="TOTAL:").font = Font(bold=True)
    ws.cell(row=row_total, column=9, value=total).font = Font(bold=True)
    ws.cell(row=row_total, column=9).number_format = '$#,##0.00'
    
    # Ajustar anchos
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except: pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)
        
    wb.save(filepath)
    return filepath
    cell_total_monto.number_format = '$#,##0.00'
    cell_total_monto.alignment = Alignment(horizontal='right')
    cell_total_monto.border = border
    
    ws.cell(row=row_num, column=10).border = border
    
    # ===== ESTADÍSTICAS =====
    row_num += 2
    ws.cell(row=row_num, column=1, value='ESTADÍSTICAS:').font = Font(bold=True)
    row_num += 1
    
    ws.cell(row=row_num, column=1, value=f"Total de recibos:")
    ws.cell(row=row_num, column=2, value=len([r for r in recibos if not r.get('eliminado')]))
    row_num += 1
    
    # Recibos por tipo
    nuevas_siembras = len([r for r in recibos if r['tipo_accion'] == 'Nueva siembra' and not r.get('eliminado')])
    riegos_adicionales = len([r for r in recibos if r['tipo_accion'] == 'Riego adicional' and not r.get('eliminado')])
    
    ws.cell(row=row_num, column=1, value=f"Nuevas siembras:")
    ws.cell(row=row_num, column=2, value=nuevas_siembras)
    row_num += 1
    
    ws.cell(row=row_num, column=1, value=f"Riegos adicionales:")
    ws.cell(row=row_num, column=2, value=riegos_adicionales)
    row_num += 1
    
    # ===== PIE DE PÁGINA =====
    row_num += 2
    ws.merge_cells(f'A{row_num}:J{row_num}')
    cell_generado = ws.cell(row=row_num, column=1)
    cell_generado.value = f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M:%S')}"
    cell_generado.font = Font(name='Calibri', size=9, italic=True, color='808080')
    cell_generado.alignment = Alignment(horizontal='center')
    
    # ===== AJUSTAR ANCHOS DE COLUMNA =====
    ws.column_dimensions['A'].width = 8      # FOLIO
    ws.column_dimensions['B'].width = 10     # NO DE LOTE
    ws.column_dimensions['C'].width = 10     # CICLO
    ws.column_dimensions['D'].width = 12     # BARRIO
    ws.column_dimensions['E'].width = 12     # CULTIVO
    ws.column_dimensions['F'].width = 14     # SUPERFICIE DE LA MILPA
    ws.column_dimensions['G'].width = 13     # RIEGO DE LA MILPA
    ws.column_dimensions['H'].width = 18     # CAMPESINO
    ws.column_dimensions['I'].width = 15     # CUANTO SE COBRA
    ws.column_dimensions['J'].width = 12     # FECHA
    
    # Guardar archivo
    wb.save(ruta_excel)
    print(f"✅ Corte de caja Excel generado: {ruta_excel}")
    return ruta_excel


def generar_pdf_estadisticas(estadisticas: Dict, estadisticas_cultivo: List[Dict]) -> str:
    """
    Genera un PDF profesional premium con las estadísticas del sistema incluyendo gráficos avanzados.
    
    Args:
        estadisticas: Diccionario con estadísticas generales
        estadisticas_cultivo: Lista de estadísticas por cultivo
    
    Returns:
        Ruta del archivo PDF generado
    """
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.units import cm, mm
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas as pdfcanvas
    from reportlab.platypus import Table, TableStyle
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import numpy as np
    import tempfile
    
    # Crear directorio si no existe
    reportes_dir = os.path.join("database", "reportes")
    os.makedirs(reportes_dir, exist_ok=True)
    
    # Nombre del archivo con fecha
    fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_archivo = f"estadisticas_{fecha_str}.pdf"
    ruta_pdf = os.path.join(reportes_dir, nombre_archivo)
    
    # Crear PDF con tamaño A4
    c = pdfcanvas.Canvas(ruta_pdf, pagesize=A4)
    ancho, alto = A4
    
    # Configuración de colores corporativos
    COLOR_PRIMARIO = colors.HexColor("#1a5490")
    COLOR_SECUNDARIO = colors.HexColor("#2e7d32")
    COLOR_ACENTO = colors.HexColor("#f57c00")
    COLOR_FONDO_CLARO = colors.HexColor("#f8f9fa")
    COLOR_FONDO_TABLA = colors.HexColor("#e8f4f8")
    
    # ===== FUNCIONES AUXILIARES PARA GRÁFICOS MEJORADOS =====
    def crear_grafico_barras_cultivos_pro(estadisticas_cultivo, ruta_imagen):
        """Crea un gráfico de barras profesional con gradientes"""
        if not estadisticas_cultivo:
            return None
        
        cultivos = [cult['cultivo'] for cult in estadisticas_cultivo]
        superficies = [cult['superficie_total'] for cult in estadisticas_cultivo]
        
        fig, ax = plt.subplots(figsize=(11, 6), facecolor='white')
        
        # Colores degradados profesionales
        colores_base = ['#2e7d32', '#1976d2', '#f57c00', '#d32f2f', '#7b1fa2', '#0097a7', '#fbc02d', '#5d4037']
        colores = colores_base[:len(cultivos)]
        
        # Crear barras con efecto visual
        barras = ax.bar(range(len(cultivos)), superficies, color=colores, 
                        edgecolor='white', linewidth=2.5, alpha=0.9)
        
        # Añadir valores sobre las barras con mejor formato
        for i, (barra, valor) in enumerate(zip(barras, superficies)):
            altura = barra.get_height()
            ax.text(barra.get_x() + barra.get_width()/2., altura + max(superficies)*0.03,
                    f'{valor:.2f} ha',
                    ha='center', va='bottom', fontsize=11, fontweight='bold',
                    bbox=dict(boxstyle='round,pad=0.4', facecolor='white', 
                            edgecolor=colores[i % len(colores)], linewidth=2, alpha=0.95))
        
        ax.set_xticks(range(len(cultivos)))
        ax.set_xticklabels(cultivos, rotation=35, ha='right', fontsize=11, fontweight='600')
        ax.set_xlabel('Tipo de Cultivo', fontsize=13, fontweight='bold', labelpad=10)
        ax.set_ylabel('Superficie (hectáreas)', fontsize=13, fontweight='bold', labelpad=10)
        ax.set_title('Distribución de Superficie por Tipo de Cultivo', 
                     fontsize=15, fontweight='bold', pad=20, color='#1a5490')
        
        # Mejorar grid
        ax.grid(axis='y', alpha=0.3, linestyle='--', linewidth=1)
        ax.set_axisbelow(True)
        
        # Añadir línea promedio
        promedio = np.mean(superficies)
        ax.axhline(y=promedio, color='red', linestyle='--', linewidth=2, alpha=0.7, 
                  label=f'Promedio: {promedio:.2f} ha')
        ax.legend(loc='upper right', fontsize=10, framealpha=0.95)
        
        # Mejorar estilo de ejes
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_linewidth(1.5)
        ax.spines['bottom'].set_linewidth(1.5)
        
        plt.tight_layout()
        plt.savefig(ruta_imagen, dpi=300, bbox_inches='tight', facecolor='white')
        plt.close()
        return ruta_imagen
    
    def crear_grafico_pastel_pro(estadisticas_cultivo, ruta_imagen):
        """Crea un gráfico de dona profesional con detalles mejorados"""
        if not estadisticas_cultivo:
            return None
        
        cultivos = [cult['cultivo'] for cult in estadisticas_cultivo]
        superficies = [cult['superficie_total'] for cult in estadisticas_cultivo]
        
        fig, ax = plt.subplots(figsize=(10, 10), facecolor='white')
        
        # Colores profesionales
        colores = ['#2e7d32', '#1976d2', '#f57c00', '#d32f2f', '#7b1fa2', '#0097a7', '#fbc02d', '#5d4037']
        colores = colores[:len(cultivos)]
        
        # Crear efecto de explosión sutil para el mayor valor
        explode = [0.08 if i == superficies.index(max(superficies)) else 0 for i in range(len(superficies))]
        
        def autopct_format(pct):
            return f'{pct:.1f}%' if pct > 2 else ''
        
        # Crear gráfico de dona
        wedges, texts, autotexts = ax.pie(superficies, labels=None, autopct=autopct_format,
                                            colors=colores, startangle=90, explode=explode,
                                            textprops={'fontsize': 12, 'fontweight': 'bold'},
                                            wedgeprops={'edgecolor': 'white', 'linewidth': 3, 
                                                       'antialiased': True},
                                            pctdistance=0.82)
        
        # Crear el agujero de la dona
        centre_circle = plt.Circle((0, 0), 0.65, fc='white', linewidth=2.5, edgecolor='#1a5490')
        ax.add_artist(centre_circle)
        
        # Texto central
        total_superficie = sum(superficies)
        ax.text(0, 0.08, f'{total_superficie:.2f}', ha='center', va='center',
                fontsize=30, fontweight='bold', color='#1a5490')
        ax.text(0, -0.18, 'hectáreas\ntotales', ha='center', va='center',
                fontsize=12, color='#666666', style='italic')
        
        # Mejorar textos de porcentaje
        for autotext in autotexts:
            autotext.set_color('white')
            autotext.set_fontweight('bold')
            autotext.set_fontsize(11)
        
        # Leyenda mejorada con superficie
        leyendas = [f'{cultivo}: {sup:.2f} ha' for cultivo, sup in zip(cultivos, superficies)]
        ax.legend(wedges, leyendas, title="Cultivos", loc="center left",
                 bbox_to_anchor=(1, 0, 0.5, 1), fontsize=11, title_fontsize=13,
                 frameon=True, shadow=True, fancybox=True)
        
        ax.set_title('Distribución Porcentual de Cultivos', 
                     fontsize=16, fontweight='bold', pad=25, color='#1a5490')
        
        plt.tight_layout()
        plt.savefig(ruta_imagen, dpi=300, bbox_inches='tight', facecolor='white')
        plt.close()
        return ruta_imagen
    
    def crear_grafico_comparativo(estadisticas_cultivo, ruta_imagen):
        """Crea un gráfico comparativo de campesinos vs superficie"""
        if not estadisticas_cultivo:
            return None
        
        cultivos = [cult['cultivo'] for cult in estadisticas_cultivo]
        campesinos = [cult['num_siembras'] for cult in estadisticas_cultivo]
        superficies = [cult['superficie_total'] for cult in estadisticas_cultivo]
        
        fig, ax1 = plt.subplots(figsize=(11, 6), facecolor='white')
        
        x = np.arange(len(cultivos))
        width = 0.38
        
        # Primer eje: Campesinos
        color1 = '#1976d2'
        ax1.bar(x - width/2, campesinos, width, label='Campesinos', 
                color=color1, alpha=0.85, edgecolor='white', linewidth=2.5)
        ax1.set_xlabel('Cultivo', fontsize=13, fontweight='bold', labelpad=10)
        ax1.set_ylabel('Número de Campesinos', color=color1, fontsize=12, fontweight='bold', labelpad=10)
        ax1.tick_params(axis='y', labelcolor=color1, labelsize=10)
        ax1.set_xticks(x)
        ax1.set_xticklabels(cultivos, rotation=35, ha='right', fontsize=10, fontweight='600')
        
        # Segundo eje: Superficie
        ax2 = ax1.twinx()
        color2 = '#2e7d32'
        ax2.bar(x + width/2, superficies, width, label='Superficie (ha)',
                color=color2, alpha=0.85, edgecolor='white', linewidth=2.5)
        ax2.set_ylabel('Superficie (hectáreas)', color=color2, fontsize=12, fontweight='bold', labelpad=10)
        ax2.tick_params(axis='y', labelcolor=color2, labelsize=10)
        
        # Título y leyenda
        plt.title('Comparativa: Campesinos vs Superficie por Cultivo',
                 fontsize=15, fontweight='bold', pad=20, color='#1a5490')
        
        # Leyenda combinada
        lines1, labels1 = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper right', 
                  fontsize=10, framealpha=0.95, shadow=True)
        
        ax1.grid(axis='y', alpha=0.3, linestyle='--', linewidth=1)
        ax1.set_axisbelow(True)
        
        fig.tight_layout()
        plt.savefig(ruta_imagen, dpi=300, bbox_inches='tight', facecolor='white')
        plt.close()
        return ruta_imagen
    
    # ===== FUNCIÓN PARA DIBUJAR ENCABEZADO PROFESIONAL =====
    def dibujar_encabezado_pagina(c, numero_pagina=1):
        """Dibuja un encabezado consistente en todas las páginas"""
        # Barra superior decorativa
        c.setFillColor(COLOR_PRIMARIO)
        c.rect(0, alto - 1*cm, ancho, 1*cm, fill=True, stroke=False)
        
        # Logo si existe
        y_logo = alto - 3.8*cm
        if os.path.exists(LOGO_PATH):
            try:
                c.drawImage(LOGO_PATH, 1.5*cm, y_logo, width=2.5*cm, height=2.5*cm, 
                           mask='auto', preserveAspectRatio=True)
            except Exception as e:
                print(f"Error al añadir logo: {e}")
        
        # Título principal
        c.setFillColor(COLOR_PRIMARIO)
        c.setFont("Helvetica-Bold", 20)
        c.drawString(4.5*cm, alto - 2.2*cm, "INFORME ESTADÍSTICO")
        
        c.setFillColor(COLOR_SECUNDARIO)
        c.setFont("Helvetica-Bold", 11)
        nombre_oficina = obtener_configuracion("nombre_oficina") or "SISTEMA DE RIEGO"
        if len(nombre_oficina) > 65:
            nombre_oficina = nombre_oficina[:62] + "..."
        c.drawString(4.5*cm, alto - 2.9*cm, nombre_oficina.upper())
        
        # Línea decorativa
        c.setStrokeColor(COLOR_ACENTO)
        c.setLineWidth(3)
        c.line(4.5*cm, alto - 3.2*cm, ancho - 1.5*cm, alto - 3.2*cm)
        
        # Fecha y número de página
        c.setFillColor(colors.HexColor("#666666"))
        c.setFont("Helvetica", 9)
        fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')
        c.drawRightString(ancho - 1.5*cm, alto - 2.2*cm, f"Fecha: {fecha_actual}")
        c.drawRightString(ancho - 1.5*cm, alto - 2.8*cm, f"Página {numero_pagina}")
        
        return alto - 4.5*cm  # Retorna posición Y inicial para contenido (más espacio)
    
    def dibujar_pie_pagina(c, numero_pagina):
        """Dibuja pie de página profesional"""
        c.setFillColor(COLOR_FONDO_CLARO)
        c.rect(0, 0, ancho, 2*cm, fill=True, stroke=False)
        
        c.setStrokeColor(COLOR_PRIMARIO)
        c.setLineWidth(2)
        c.line(1.5*cm, 1.8*cm, ancho - 1.5*cm, 1.8*cm)
        
        c.setFont("Helvetica-Oblique", 8)
        c.setFillColor(colors.HexColor("#666666"))
        c.drawCentredString(ancho/2, 1*cm, 
            f"Documento generado automáticamente el {datetime.now().strftime('%d/%m/%Y a las %H:%M:%S')}")
        c.drawCentredString(ancho/2, 0.5*cm, "Sistema de Gestión Agrícola - Todos los derechos reservados")
    
    # ===== PÁGINA 1: RESUMEN EJECUTIVO =====
    ypos = dibujar_encabezado_pagina(c, 1)
    
    # Sección de resumen ejecutivo
    c.setFillColor(COLOR_PRIMARIO)
    c.roundRect(1.5*cm, ypos - 0.8*cm, ancho - 3*cm, 0.8*cm, 0.3*cm, fill=True, stroke=False)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 13)
    c.drawCentredString(ancho/2, ypos - 0.5*cm, "📊 RESUMEN EJECUTIVO")
    
    ypos -= 5*cm
    
    # Tabla de resumen con diseño mejorado
    datos_generales = [
        ["INDICADOR", "VALOR"],
        ["Total de Campesinos Registrados", str(estadisticas.get("total_campesinos", 0))],
        ["Total de Lotes Asignados", str(estadisticas.get("total_lotes", 0))],
        ["Superficie Total Disponible", f"{estadisticas.get('superficie_total', 0):.2f} ha"],
        ["Hectáreas Sembradas", f"{estadisticas.get('hectareas_sembradas', 0):.2f} ha"],
        ["Hectáreas Sin Cultivar", f"{estadisticas.get('hectareas_sin_sembrar', 0):.2f} ha"],
        ["Porcentaje de Ocupación", f"{estadisticas.get('porcentaje_sembrado', 0):.1f}%"],
        ["Siembras Activas", str(estadisticas.get("siembras_activas", 0))],
        ["Campesinos Sin Siembra", str(estadisticas.get("campesinos_sin_siembra", 0))],
        ["Total de Recibos Emitidos", str(estadisticas.get("total_recibos", 0))],
        ["Ingresos Totales", f"${estadisticas.get('ingresos_totales', 0):,.2f} MXN"],
    ]
    
    tabla_general = Table(datos_generales, colWidths=[11*cm, 5.5*cm])
    tabla_general.setStyle(TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), COLOR_PRIMARIO),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        # Contenido
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('LEFTPADDING', (0, 1), (-1, -1), 12),
        ('RIGHTPADDING', (0, 1), (-1, -1), 12),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        # Bordes y colores alternos
        ('GRID', (0, 0), (-1, -1), 1.5, COLOR_PRIMARIO),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, COLOR_FONDO_TABLA]),
        # Resaltar última fila
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#fff3e0")),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (1, -1), (1, -1), COLOR_ACENTO),
    ]))
    
    altura_tabla = len(datos_generales) * 0.7 * cm
    tabla_general.wrapOn(c, ancho, alto)
    tabla_general.drawOn(c, 1.5*cm, ypos - altura_tabla)
    ypos -= altura_tabla + 1.2*cm
    
    # Indicadores clave en tarjetas (más espaciadas)
    if estadisticas.get('porcentaje_sembrado', 0) > 0:
        # Tarjeta de eficiencia
        c.setFillColor(colors.HexColor("#e8f5e9"))
        c.roundRect(1.5*cm, ypos - 2.5*cm, 8*cm, 2.5*cm, 0.4*cm, fill=True, stroke=False)
        c.setStrokeColor(COLOR_SECUNDARIO)
        c.setLineWidth(3)
        c.roundRect(1.5*cm, ypos - 2.5*cm, 8*cm, 2.5*cm, 0.4*cm, fill=False, stroke=True)
        
        c.setFillColor(COLOR_SECUNDARIO)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(2*cm, ypos - 0.9*cm, "📈 EFICIENCIA DE USO")
        c.setFont("Helvetica-Bold", 26)
        c.drawString(2*cm, ypos - 1.6*cm, f"{estadisticas.get('porcentaje_sembrado', 0):.1f}%")
        c.setFont("Helvetica", 9)
        c.setFillColor(colors.HexColor("#666666"))
        c.drawString(2*cm, ypos - 2.1*cm, "de la superficie está cultivada")
        
        # Tarjeta de productividad
        c.setFillColor(colors.HexColor("#e3f2fd"))
        c.roundRect(10*cm, ypos - 2.5*cm, 8*cm, 2.5*cm, 0.4*cm, fill=True, stroke=False)
        c.setStrokeColor(colors.HexColor("#1976d2"))
        c.setLineWidth(3)
        c.roundRect(10*cm, ypos - 2.5*cm, 8*cm, 2.5*cm, 0.4*cm, fill=False, stroke=True)
        
        campesinos_activos = estadisticas.get("total_campesinos", 0) - estadisticas.get("campesinos_sin_siembra", 0)
        tasa_participacion = (campesinos_activos / estadisticas.get("total_campesinos", 1)) * 100 if estadisticas.get("total_campesinos", 0) > 0 else 0
        
        c.setFillColor(colors.HexColor("#1976d2"))
        c.setFont("Helvetica-Bold", 11)
        c.drawString(10.5*cm, ypos - 0.9*cm, "👨‍🌾 TASA DE PARTICIPACIÓN")
        c.setFont("Helvetica-Bold", 26)
        c.drawString(10.5*cm, ypos - 1.6*cm, f"{tasa_participacion:.1f}%")
        c.setFont("Helvetica", 9)
        c.setFillColor(colors.HexColor("#666666"))
        c.drawString(10.5*cm, ypos - 2.1*cm, f"{campesinos_activos} campesinos activos")
    
    dibujar_pie_pagina(c, 1)
    
    # ===== PÁGINA 2: GRÁFICO DE BARRAS =====
    c.showPage()
    ypos = dibujar_encabezado_pagina(c, 2)
    
    if estadisticas_cultivo:
        c.setFillColor(COLOR_SECUNDARIO)
        c.roundRect(1.5*cm, ypos - 0.8*cm, ancho - 3*cm, 0.8*cm, 0.3*cm, fill=True, stroke=False)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 13)
        c.drawCentredString(ancho/2, ypos - 0.5*cm, "📊 ANÁLISIS DE SUPERFICIE POR CULTIVO")
        
        ypos -= 1.8*cm
        
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
            ruta_grafico_barras = tmp.name
        
        crear_grafico_barras_cultivos_pro(estadisticas_cultivo, ruta_grafico_barras)
        
        try:
            c.drawImage(ruta_grafico_barras, 1.5*cm, ypos - 14*cm, 
                       width=ancho - 3*cm, height=13.5*cm, preserveAspectRatio=True, mask='auto')
        except Exception as e:
            print(f"Error al insertar gráfico de barras: {e}")
        finally:
            try:
                os.remove(ruta_grafico_barras)
            except:
                pass
    
    dibujar_pie_pagina(c, 2)
    
    # ===== PÁGINA 3: GRÁFICO DE DONA =====
    c.showPage()
    ypos = dibujar_encabezado_pagina(c, 3)
    
    if estadisticas_cultivo:
        c.setFillColor(COLOR_ACENTO)
        c.roundRect(1.5*cm, ypos - 0.8*cm, ancho - 3*cm, 0.8*cm, 0.3*cm, fill=True, stroke=False)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 13)
        c.drawCentredString(ancho/2, ypos - 0.5*cm, "🥧 DISTRIBUCIÓN PORCENTUAL DE CULTIVOS")
        
        ypos -= 1.8*cm
        
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
            ruta_grafico_pastel = tmp.name
        
        crear_grafico_pastel_pro(estadisticas_cultivo, ruta_grafico_pastel)
        
        try:
            c.drawImage(ruta_grafico_pastel, 0.8*cm, ypos - 15*cm,
                       width=ancho - 1.6*cm, height=14.5*cm, preserveAspectRatio=True, mask='auto')
        except Exception as e:
            print(f"Error al insertar gráfico de pastel: {e}")
        finally:
            try:
                os.remove(ruta_grafico_pastel)
            except:
                pass
    
    dibujar_pie_pagina(c, 3)
    
    # ===== PÁGINA 4: GRÁFICO COMPARATIVO =====
    c.showPage()
    ypos = dibujar_encabezado_pagina(c, 4)
    
    if estadisticas_cultivo:
        c.setFillColor(colors.HexColor("#7b1fa2"))
        c.roundRect(1.5*cm, ypos - 0.8*cm, ancho - 3*cm, 0.8*cm, 0.3*cm, fill=True, stroke=False)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 13)
        c.drawCentredString(ancho/2, ypos - 0.5*cm, "📈 COMPARATIVA: CAMPESINOS VS SUPERFICIE")
        
        ypos -= 1.8*cm
        
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
            ruta_grafico_comparativo = tmp.name
        
        crear_grafico_comparativo(estadisticas_cultivo, ruta_grafico_comparativo)
        
        try:
            c.drawImage(ruta_grafico_comparativo, 1.5*cm, ypos - 14*cm,
                       width=ancho - 3*cm, height=13.5*cm, preserveAspectRatio=True, mask='auto')
        except Exception as e:
            print(f"Error al insertar gráfico comparativo: {e}")
        finally:
            try:
                os.remove(ruta_grafico_comparativo)
            except:
                pass
    
    dibujar_pie_pagina(c, 4)
    
    # ===== PÁGINA 5: TABLA DETALLADA =====
    c.showPage()
    ypos = dibujar_encabezado_pagina(c, 5)
    
    c.setFillColor(COLOR_PRIMARIO)
    c.roundRect(1.5*cm, ypos - 0.8*cm, ancho - 3*cm, 0.8*cm, 0.3*cm, fill=True, stroke=False)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 13)
    c.drawCentredString(ancho/2, ypos - 0.5*cm, "🌾 DETALLE COMPLETO POR CULTIVO")
    
    ypos -= 5*cm
    
    if estadisticas_cultivo:
        datos_cultivos = [["CULTIVO", "CAMPESINOS", "SUPERFICIE\n(ha)", "% DEL\nTOTAL", "PROMEDIO/\nCAMPESINO"]]
        total_superficie = estadisticas.get('hectareas_sembradas', 0)
        
        for cultivo in estadisticas_cultivo:
            sup = cultivo['superficie_total']
            num = cultivo['num_siembras']
            porcentaje = (sup / total_superficie * 100) if total_superficie > 0 else 0
            promedio = sup / num if num > 0 else 0
            datos_cultivos.append([
                cultivo["cultivo"],
                str(num),
                f"{sup:.2f}",
                f"{porcentaje:.1f}%",
                f"{promedio:.2f} ha"
            ])
        
        # Agregar fila de totales
        total_campesinos = sum(c['num_siembras'] for c in estadisticas_cultivo)
        datos_cultivos.append([
            "TOTAL",
            str(total_campesinos),
            f"{total_superficie:.2f}",
            "100.0%",
            f"{total_superficie/total_campesinos if total_campesinos > 0 else 0:.2f} ha"
        ])
        
        tabla_cultivos = Table(datos_cultivos, colWidths=[4.2*cm, 3*cm, 3.2*cm, 3*cm, 3.6*cm])
        tabla_cultivos.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), COLOR_PRIMARIO),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            # Contenido
            ('BACKGROUND', (0, 1), (-1, -2), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
            # Colores alternos
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, COLOR_FONDO_TABLA]),
            # Fila de totales
            ('BACKGROUND', (0, -1), (-1, -1), COLOR_SECUNDARIO),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
            ('TOPPADDING', (0, -1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 10),
            # Bordes
            ('GRID', (0, 0), (-1, -1), 1.5, COLOR_PRIMARIO),
            ('LINEBELOW', (0, 0), (-1, 0), 2.5, COLOR_PRIMARIO),
            ('LINEABOVE', (0, -1), (-1, -1), 2.5, COLOR_SECUNDARIO),
        ]))
        
        altura_tabla = len(datos_cultivos) * 0.7 * cm
        tabla_cultivos.wrapOn(c, ancho, alto)
        tabla_cultivos.drawOn(c, 1.5*cm, ypos - altura_tabla)
        
        ypos -= altura_tabla + 1.5*cm
        
        # Agregar notas y observaciones
        if ypos > 6*cm:
            c.setFillColor(colors.HexColor("#fff3e0"))
            c.roundRect(1.5*cm, ypos - 3*cm, ancho - 3*cm, 3*cm, 0.4*cm, fill=True, stroke=False)
            
            c.setStrokeColor(COLOR_ACENTO)
            c.setLineWidth(2.5)
            c.roundRect(1.5*cm, ypos - 3*cm, ancho - 3*cm, 3*cm, 0.4*cm, fill=False, stroke=True)
            
            c.setFillColor(COLOR_ACENTO)
            c.setFont("Helvetica-Bold", 11)
            c.drawString(2*cm, ypos - 0.8*cm, "📋 NOTAS Y OBSERVACIONES")
            
            c.setFillColor(colors.black)
            c.setFont("Helvetica", 9)
            
            # Calcular estadística relevante
            if estadisticas_cultivo:
                cultivo_mayor = max(estadisticas_cultivo, key=lambda x: x['superficie_total'])
                cultivo_menor = min(estadisticas_cultivo, key=lambda x: x['superficie_total'])
                
                notas = [
                    f"• El cultivo con mayor superficie es {cultivo_mayor['cultivo']} con {cultivo_mayor['superficie_total']:.2f} ha",
                    f"• El cultivo con menor superficie es {cultivo_menor['cultivo']} con {cultivo_menor['superficie_total']:.2f} ha",
                    f"• Promedio de superficie por campesino: {total_superficie/total_campesinos if total_campesinos > 0 else 0:.2f} hectáreas",
                    f"• Se registran {len(estadisticas_cultivo)} tipos diferentes de cultivos en el sistema"
                ]
                
                y_nota = ypos - 1.3*cm
                for nota in notas:
                    c.drawString(2.2*cm, y_nota, nota)
                    y_nota -= 0.5*cm
    
    dibujar_pie_pagina(c, 5)
    
    # ===== PÁGINA 6: ANÁLISIS Y RECOMENDACIONES =====
    c.showPage()
    ypos = dibujar_encabezado_pagina(c, 6)
    
    c.setFillColor(colors.HexColor("#d32f2f"))
    c.roundRect(1.5*cm, ypos - 0.8*cm, ancho - 3*cm, 0.8*cm, 0.3*cm, fill=True, stroke=False)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 13)
    c.drawCentredString(ancho/2, ypos - 0.5*cm, "💡 ANÁLISIS Y RECOMENDACIONES")
    
    ypos -= 1.8*cm
    
    # Sección de análisis (más espaciada)
    c.setFillColor(colors.HexColor("#e8f5e9"))
    c.roundRect(1.5*cm, ypos - 7*cm, ancho - 3*cm, 7*cm, 0.4*cm, fill=True, stroke=False)
    
    c.setStrokeColor(COLOR_SECUNDARIO)
    c.setLineWidth(2.5)
    c.roundRect(1.5*cm, ypos - 7*cm, ancho - 3*cm, 7*cm, 0.4*cm, fill=False, stroke=True)
    
    c.setFillColor(COLOR_SECUNDARIO)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, ypos - 0.7*cm, "✅ PUNTOS FUERTES")
    
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 10)
    
    puntos_fuertes = []
    porcentaje_sembrado = estadisticas.get('porcentaje_sembrado', 0)
    
    if porcentaje_sembrado >= 85:
        puntos_fuertes.append(f"• Excelente tasa de utilización: {porcentaje_sembrado:.1f}% de superficie cultivada")
    elif porcentaje_sembrado >= 70:
        puntos_fuertes.append(f"• Buena tasa de utilización: {porcentaje_sembrado:.1f}% de superficie cultivada")
    
    if estadisticas.get('total_campesinos', 0) > 0:
        tasa_activos = ((estadisticas.get('total_campesinos', 0) - estadisticas.get('campesinos_sin_siembra', 0)) / estadisticas.get('total_campesinos', 1)) * 100
        if tasa_activos >= 70:
            puntos_fuertes.append(f"• Alta participación: {tasa_activos:.1f}% de campesinos con siembras activas")
    
    if estadisticas_cultivo:
        puntos_fuertes.append(f"• Diversificación: {len(estadisticas_cultivo)} tipos diferentes de cultivos registrados")
    
    if estadisticas.get('ingresos_totales', 0) > 0:
        puntos_fuertes.append(f"• Ingresos documentados: ${estadisticas.get('ingresos_totales', 0):,.2f} MXN registrados")
    
    y_texto = ypos - 1.3*cm
    for punto in puntos_fuertes:
        c.drawString(2.2*cm, y_texto, punto)
        y_texto -= 0.6*cm
    
    # Áreas de oportunidad
    y_texto -= 0.5*cm
    c.setFillColor(COLOR_ACENTO)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, y_texto, "⚠️ ÁREAS DE OPORTUNIDAD")
    
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 10)
    
    oportunidades = []
    
    if porcentaje_sembrado < 85:
        hectareas_disponibles = estadisticas.get('hectareas_sin_sembrar', 0)
        oportunidades.append(f"• Potencial de expansión: {hectareas_disponibles:.2f} ha disponibles para cultivo")
    
    if estadisticas.get('campesinos_sin_siembra', 0) > 0:
        oportunidades.append(f"• {estadisticas.get('campesinos_sin_siembra', 0)} campesinos sin siembra activa podrían integrarse")
    
    if estadisticas_cultivo:
        superficies = [c['superficie_total'] for c in estadisticas_cultivo]
        max_sup = max(superficies)
        min_sup = min(superficies)
        if max_sup > min_sup * 2:
            oportunidades.append("• Considerar equilibrar distribución entre cultivos menos representados")
    
    y_texto -= 0.6*cm
    for oportunidad in oportunidades:
        c.drawString(2.2*cm, y_texto, oportunidad)
        y_texto -= 0.6*cm
    
    ypos -= 8*cm
    
    # Recomendaciones estratégicas
    c.setFillColor(colors.HexColor("#e3f2fd"))
    c.roundRect(1.5*cm, ypos - 6*cm, ancho - 3*cm, 6*cm, 0.4*cm, fill=True, stroke=False)
    
    c.setStrokeColor(colors.HexColor("#1976d2"))
    c.setLineWidth(2.5)
    c.roundRect(1.5*cm, ypos - 6*cm, ancho - 3*cm, 6*cm, 0.4*cm, fill=False, stroke=True)
    
    c.setFillColor(colors.HexColor("#1976d2"))
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, ypos - 0.7*cm, "🎯 RECOMENDACIONES ESTRATÉGICAS")
    
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 10)
    
    recomendaciones = [
        "1. Implementar programas de capacitación para campesinos sin siembra activa",
        "2. Promover cultivos con mayor rentabilidad en áreas disponibles",
        "3. Establecer sistema de seguimiento trimestral de indicadores clave",
        "4. Considerar rotación de cultivos para optimizar el uso del suelo",
        "5. Fortalecer el registro y documentación de operaciones financieras",
        "6. Evaluar implementación de tecnologías de riego más eficientes"
    ]
    
    y_texto = ypos - 1.3*cm
    for rec in recomendaciones:
        c.drawString(2.2*cm, y_texto, rec)
        y_texto -= 0.6*cm
    
    # Conclusión
    ypos -= 7*cm
    
    if ypos > 5*cm:
        c.setFillColor(colors.HexColor("#f3e5f5"))
        c.roundRect(1.5*cm, ypos - 3.5*cm, ancho - 3*cm, 3.5*cm, 0.4*cm, fill=True, stroke=False)
        
        c.setStrokeColor(colors.HexColor("#7b1fa2"))
        c.setLineWidth(2.5)
        c.roundRect(1.5*cm, ypos - 3.5*cm, ancho - 3*cm, 3.5*cm, 0.4*cm, fill=False, stroke=True)
        
        c.setFillColor(colors.HexColor("#7b1fa2"))
        c.setFont("Helvetica-Bold", 12)
        c.drawString(2*cm, ypos - 0.7*cm, "📝 CONCLUSIÓN")
        
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 10)
        
        conclusion_texto = f"""El sistema presenta un nivel de operación {'excelente' if porcentaje_sembrado >= 85 else 'satisfactorio' if porcentaje_sembrado >= 70 else 'con potencial'} 
        con {estadisticas.get('total_campesinos', 0)} campesinos registrados y {estadisticas.get('hectareas_sembradas', 0):.2f} hectáreas
        en producción. La diversificación de {len(estadisticas_cultivo) if estadisticas_cultivo else 0} tipos de cultivos indica una estrategia
        agrícola balanceada. Se recomienda mantener el monitoreo continuo y aplicar las
        recomendaciones estratégicas propuestas para optimizar los resultados."""
        
        y_texto = ypos - 1.3*cm
        for linea in conclusion_texto.split('\n'):
            c.drawString(2.2*cm, y_texto, linea.strip())
            y_texto -= 0.5*cm
    
    dibujar_pie_pagina(c, 6)
    
    # Guardar PDF
    c.save()
    print(f"✓ PDF profesional de estadísticas generado exitosamente: {ruta_pdf}")
    print(f"  📄 Páginas: 6")
    print(f"  📊 Gráficos: 3 (Barras, Dona, Comparativo)")
    print(f"  📈 Análisis completo incluido")
    return ruta_pdf

def generar_pdf_auditoria(registros_auditoria: List[Dict], fecha_inicio=None, fecha_fin=None) -> str:
    """
    Genera un PDF profesional con el historial de auditoría.
    
    Args:
        registros_auditoria: Lista de registros de auditoría
        fecha_inicio: Fecha de inicio del rango (opcional)
        fecha_fin: Fecha de fin del rango (opcional)
    
    Returns:
        Ruta del archivo PDF generado
    """
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas as pdf_canvas
    from reportlab.platypus import Table, TableStyle, PageBreak
    
    # Crear directorio si no existe
    reportes_dir = os.path.join('database', 'reportes')
    os.makedirs(reportes_dir, exist_ok=True)
    
    # Nombre del archivo
    fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f"auditoria_{fecha_str}.pdf"
    ruta_pdf = os.path.join(reportes_dir, nombre_archivo)
    
    # Usar orientación horizontal para más espacio
    c = pdf_canvas.Canvas(ruta_pdf, pagesize=landscape(letter))
    ancho, alto = landscape(letter)
    
    # ===== FUNCIÓN PARA DIBUJAR ENCABEZADO =====
    def dibujar_encabezado(c, y_pos):
        # Logo si existe
        if os.path.exists(LOGO_PATH):
            try:
                c.drawImage(LOGO_PATH, 2*cm, y_pos - 1.5*cm, width=1.5*cm, height=1.5*cm, mask='auto')
            except:
                pass
        
        # Título
        c.setFont("Helvetica-Bold", 16)
        c.drawString(4.5*cm, y_pos, "REGISTRO DE AUDITORÍA")
        y_pos -= 0.5*cm
        
        nombre_oficina = obtener_configuracion('nombre_oficina') or 'SISTEMA DE RIEGO'
        c.setFont("Helvetica", 10)
        c.drawString(4.5*cm, y_pos, nombre_oficina)
        y_pos -= 0.3*cm
        
        if fecha_inicio and fecha_fin:
            c.drawString(4.5*cm, y_pos, f"Período: {fecha_inicio} - {fecha_fin}")
            y_pos -= 0.3*cm
        
        c.setFont("Helvetica", 9)
        c.drawString(4.5*cm, y_pos, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        y_pos -= 0.5*cm
        
        c.line(2*cm, y_pos, ancho - 2*cm, y_pos)
        y_pos -= 0.5*cm
        
        return y_pos
    
    # Primera página
    pagina = 1
    y_pos = alto - 1.5*cm
    y_pos = dibujar_encabezado(c, y_pos)
    
    # ===== RESUMEN =====
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, y_pos, f"📋 RESUMEN: {len(registros_auditoria)} registros totales")
    y_pos -= 1*cm
    
    # ===== TABLA DE REGISTROS =====
    registros_por_pagina = 20
    
    for i in range(0, len(registros_auditoria), registros_por_pagina):
        if i > 0:
            # Nueva página
            c.showPage()
            pagina += 1
            y_pos = alto - 1.5*cm
            y_pos = dibujar_encabezado(c, y_pos)
        
        chunk = registros_auditoria[i:i+registros_por_pagina]
        
        datos_tabla = [['Fecha/Hora', 'Tipo', 'Descripción', 'Usuario/Lote']]
        
        for reg in chunk:
            fecha_hora = f"{reg['fecha']} {reg['hora']}"
            tipo = reg['tipo_accion']
            descripcion = reg['descripcion'][:50] + '...' if len(reg['descripcion']) > 50 else reg['descripcion']
            usuario = f"ID: {reg['campesino_id']}" if reg['campesino_id'] else '-'
            
            datos_tabla.append([fecha_hora, tipo, descripcion, usuario])
        
        tabla = Table(datos_tabla, colWidths=[4.5*cm, 4*cm, 10*cm, 3*cm])
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F497D')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        
        altura_tabla = (len(datos_tabla) + 1) * 0.5 * cm
        tabla.wrapOn(c, ancho, alto)
        tabla.drawOn(c, 2*cm, y_pos - altura_tabla)
        
        # Pie de página
        c.setFont("Helvetica-Oblique", 8)
        c.setFillColor(colors.grey)
        c.drawCentredString(ancho/2, 1*cm, f"Página {pagina} - Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M:%S')}")
        c.setFillColor(colors.black)
    
    c.save()
    
    print(f"✅ PDF de auditoría generado: {ruta_pdf}")
    return ruta_pdf

def generar_recibo_cuota_pdf_temporal(recibo_cuota_id: int) -> str:
    """Genera el PDF temporal de un recibo de cuota de cooperación"""
    from modules.cuotas import obtener_recibo_cuota
    
    recibo = obtener_recibo_cuota(recibo_cuota_id)
    
    if not recibo:
        raise ValueError("Recibo de cuota no encontrado")
    
    nombre_oficina = obtener_configuracion('nombre_oficina') or "ASOCIACIÓN DE RIEGO"
    ubicacion = obtener_configuracion('ubicacion') or "Tezontepec de Aldama, Hgo."
    
    # Crear carpeta temporal
    if platform.system() == 'Windows':
        tempdir = os.path.join(os.environ.get('TEMP', os.getcwd()), 'recibos_temp')
    else:
        tempdir = os.path.join('/tmp', 'recibos_temp')
    
    os.makedirs(tempdir, exist_ok=True)
    
    filename = f"cuota_{recibo['folio']}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    filepath = os.path.join(tempdir, filename)
    
    c = canvas.Canvas(filepath, pagesize=(RECIBO_ANCHO, RECIBO_ALTO))
    
    dibujar_recibo_cuota(c, recibo, nombre_oficina, ubicacion)
    
    c.save()
    
    return filepath


def dibujar_recibo_cuota(c, recibo: Dict, nombre_oficina: str, ubicacion: str):
    """Dibuja el recibo de cuota de cooperación (similar al diseño de riegos)"""
    
    # COLORES
    COLOR_VERDE = colors.HexColor('#B8D1BF')
    COLOR_BEIGE = colors.HexColor('#FFFFFF')
    COLOR_BEIGE_OSCURO = colors.HexColor('#C9B99A')
    COLOR_TEXTO = colors.HexColor('#2C3E2E')
    COLOR_TEXTO_GRIS = colors.HexColor('#666666')
    
    # FONDO BEIGE
    c.setFillColor(COLOR_BEIGE)
    c.roundRect(0.15*cm, 0.15*cm, RECIBO_ANCHO - 0.3*cm, RECIBO_ALTO - 0.3*cm, 0.5*cm, stroke=0, fill=1)
    
    # HEADER VERDE
    c.setFillColor(COLOR_VERDE)
    c.roundRect(0.4*cm, RECIBO_ALTO - 2.3*cm, RECIBO_ANCHO - 0.8*cm, 1.9*cm, 0.4*cm, stroke=0, fill=1)
    
    # LOGO
    if os.path.exists(LOGO_PATH):
        try:
            c.drawImage(LOGO_PATH, 0.7*cm, RECIBO_ALTO - 2.1*cm, width=1.7*cm, height=1.7*cm, mask='auto')
        except:
            pass
    
    # TÍTULO
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 10)
    c.drawCentredString((RECIBO_ANCHO/2) + 0.5*cm, RECIBO_ALTO - 1*cm, "ASOCIACIÓN DE CAMPESINOS DE BOMBEO Y REBOMBEO")
    c.setFont("Helvetica-Bold", 9)
    c.drawCentredString((RECIBO_ANCHO/2) + 0.5*cm, RECIBO_ALTO - 1.35*cm, "DEL CERRO DEL XICUCO A.C. (M7-1)")
    c.setFont("Helvetica", 8)
    c.drawCentredString((RECIBO_ANCHO/2) + 0.5*cm, RECIBO_ALTO - 1.75*cm, "RFC: ACB030619G68")
    
    # SEPARADOR
    ypos = RECIBO_ALTO - 2.45*cm
    c.setStrokeColor(COLOR_BEIGE_OSCURO)
    c.setLineWidth(0.5)
    c.line(0.7*cm, ypos, RECIBO_ANCHO - 0.7*cm, ypos)
    
    # CAJA DE DATOS
    ypos -= 0.2*cm
    c.setFillColor(colors.white)
    c.roundRect(0.7*cm, ypos - 1.5*cm, RECIBO_ANCHO - 1.4*cm, 1.4*cm, 0.25*cm, stroke=1, fill=1)
    
    # DATOS - GRID
    c.setFillColor(COLOR_TEXTO)
    c.setFont("Helvetica-Bold", 8.5)
    
    col1 = 1*cm
    col2 = 6.5*cm
    col3 = 12*cm
    col4 = 17.5*cm
    
    row_y = ypos - 0.45*cm
    
    # FILA 1
    c.drawString(col1, row_y, f"NO. RECIBO: {recibo['folio']}")
    c.drawString(col2, row_y, f"No. Lote: {recibo['numero_lote']}")
    c.drawString(col3, row_y, f"Barrio: {recibo['barrio']}")
    
    # FILA 2
    row_y -= 0.45*cm
    c.drawString(col1, row_y, f"Cuota: {recibo['nombre_cuota']}")
    
    # RECIBÍ DE
    ypos = row_y - 0.95*cm
    c.setFillColor(COLOR_TEXTO_GRIS)
    c.setFont("Helvetica", 8.5)
    c.drawString(0.8*cm, ypos, "Recibí de:")
    
    c.setFillColor(COLOR_TEXTO)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2.2*cm, ypos, recibo['nombre_campesino'].upper())
    
    # CONCEPTO
    ypos -= 0.5*cm
    c.setStrokeColor(COLOR_BEIGE_OSCURO)
    c.line(0.8*cm, ypos, RECIBO_ANCHO - 0.8*cm, ypos)
    
    ypos -= 0.3*cm
    c.setFillColor(COLOR_TEXTO_GRIS)
    c.setFont("Helvetica", 8)
    c.drawString(0.8*cm, ypos, f"Concepto: {recibo['nombre_cuota']}")
    
    # TOTAL
    ypos -= 0.5*cm
    c.setFillColor(COLOR_TEXTO)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(0.8*cm, ypos, "TOTAL:")
    
    # Caja del monto
    monto_x = RECIBO_ANCHO - 5*cm
    c.setFillColor(colors.white)
    c.setStrokeColor(COLOR_VERDE)
    c.setLineWidth(1.5)
    c.roundRect(monto_x, ypos - 0.35*cm, 4.2*cm, 0.75*cm, 0.25*cm, stroke=1, fill=1)
    
    c.setFillColor(COLOR_TEXTO_GRIS)
    c.setFont("Helvetica", 6.5)
    c.drawString(monto_x + 0.2*cm, ypos + 0.15*cm, "pago en efectivo")
    
    c.setFillColor(COLOR_TEXTO)
    c.setFont("Helvetica-Bold", 15)
    c.drawRightString(monto_x + 4*cm, ypos - 0.15*cm, f"${recibo['monto']:.2f}")
    
    # FOOTER
    ypos -= 0.9*cm
    c.setStrokeColor(COLOR_BEIGE_OSCURO)
    c.setLineWidth(0.5)
    c.line(0.8*cm, ypos, RECIBO_ANCHO - 0.8*cm, ypos)
    
    ypos -= 0.3*cm
    c.setFillColor(COLOR_TEXTO_GRIS)
    c.setFont("Helvetica", 7.5)
    
    fecha_obj = datetime.strptime(recibo['fecha'], '%Y-%m-%d')
    c.drawString(0.8*cm, ypos, f"C. Juan Aldama #25, Col. Centro, Tezontepec de Aldama. Fecha: {fecha_obj.strftime('%d/%m/%Y')}")
    
    ypos -= 0.28*cm
    hora_obj = datetime.strptime(recibo['hora'], '%H:%M:%S')
    ampm = "p.m." if hora_obj.hour >= 12 else "a.m."
    hora_12 = hora_obj.hour if hora_obj.hour <= 12 else hora_obj.hour - 12
    if hora_12 == 0:
        hora_12 = 12
    
    c.drawString(0.8*cm, ypos, f"Hora: {hora_12:02d}:{hora_obj.minute:02d}:{hora_obj.second:02d} {ampm}")
    
    # Firma
    c.drawRightString(RECIBO_ANCHO - 0.8*cm, ypos + 0.28*cm, "Firma Recaudador")
    c.line(RECIBO_ANCHO - 4*cm, ypos + 0.18*cm, RECIBO_ANCHO - 0.8*cm, ypos + 0.18*cm)
    
    # LEYENDA LEGAL
    ypos -= 0.45*cm
    c.setFont("Helvetica", 6)
    c.drawString(0.7*cm, ypos, "Este recibo ampara el pago de cuota de cooperación destinada exclusivamente al mantenimiento y operación del sistema de riego.")
    ypos -= 0.22*cm
    c.drawString(0.7*cm, ypos, "Exento de IVA conforme al régimen de personas morales con fines no lucrativos (Art. 79-80 LISR y Art. 15 fracc. XII LIVA).")


def generar_reporte_cuota_pdf(tipo_cuota_id: int) -> str:
    """Genera un reporte PDF completo de una cuota específica"""
    from modules.cuotas import get_cuotas_connection, obtener_resumen_cuota
    
    conn = get_cuotas_connection()
    cursor = conn.cursor()
    
    # Obtener información de la cuota
    cursor.execute("SELECT * FROM tipos_cuota WHERE id = ?", (tipo_cuota_id,))
    cuota = dict(cursor.fetchone())
    
    # Obtener todos los campesinos asignados
    cursor.execute("""
        SELECT * FROM cuotas_campesinos
        WHERE tipo_cuota_id = ?
        ORDER BY pagado ASC, numero_lote ASC
    """, (tipo_cuota_id,))
    
    campesinos = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    # Obtener resumen
    resumen = obtener_resumen_cuota(tipo_cuota_id)
    
    # Crear PDF
    reportes_dir = os.path.join('database', 'reportes')
    os.makedirs(reportes_dir, exist_ok=True)
    
    fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f"reporte_cuota_{cuota['nombre'].replace(' ', '_')}_{fecha_str}.pdf"
    ruta_pdf = os.path.join(reportes_dir, nombre_archivo)
    
    c = canvas.Canvas(ruta_pdf, pagesize=letter)
    ancho, alto = letter
    
    # LOGO
    if os.path.exists(LOGO_PATH):
        try:
            c.drawImage(LOGO_PATH, 2*cm, alto - 3*cm, width=2*cm, height=2*cm, mask='auto')
        except:
            pass
    
    # ENCABEZADO
    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(ancho/2, alto - 2*cm, "REPORTE DE CUOTA DE COOPERACIÓN")
    
    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(ancho/2, alto - 2.8*cm, cuota['nombre'].upper())
    
    c.setFont("Helvetica", 10)
    c.drawCentredString(ancho/2, alto - 3.3*cm, f"Monto: ${cuota['monto']:.2f}")
    c.drawCentredString(ancho/2, alto - 3.8*cm, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    
    # RESUMEN
    ypos = alto - 5*cm
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, ypos, "RESUMEN")
    
    ypos -= 0.8*cm
    c.setFont("Helvetica", 10)
    c.drawString(2*cm, ypos, f"Total Asignados: {resumen['total_asignados']}")
    ypos -= 0.5*cm
    c.drawString(2*cm, ypos, f"Total Pagados: {resumen['total_pagados']} - Monto Recaudado: ${resumen['monto_recaudado']:.2f}")
    ypos -= 0.5*cm
    c.drawString(2*cm, ypos, f"Total Pendientes: {resumen['total_pendientes']} - Monto Pendiente: ${resumen['monto_pendiente']:.2f}")
    
    # TABLA
    ypos -= 1*cm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2*cm, ypos, "DETALLE DE CAMPESINOS")
    
    ypos -= 0.6*cm
    
    # Encabezados de tabla
    c.setFont("Helvetica-Bold", 8)
    c.drawString(2*cm, ypos, "Lote")
    c.drawString(3.5*cm, ypos, "Nombre")
    c.drawString(11*cm, ypos, "Barrio")
    c.drawString(13.5*cm, ypos, "Monto")
    c.drawString(15.5*cm, ypos, "Estado")
    c.drawString(17.5*cm, ypos, "Fecha Pago")
    
    ypos -= 0.3*cm
    c.line(2*cm, ypos, ancho - 2*cm, ypos)
    ypos -= 0.4*cm
    
    # Datos
    c.setFont("Helvetica", 7)
    
    for campesino in campesinos:
        if ypos < 3*cm:  # Nueva página si se acaba el espacio
            c.showPage()
            ypos = alto - 2*cm
            c.setFont("Helvetica-Bold", 8)
            c.drawString(2*cm, ypos, "Lote")
            c.drawString(3.5*cm, ypos, "Nombre")
            c.drawString(11*cm, ypos, "Barrio")
            c.drawString(13.5*cm, ypos, "Monto")
            c.drawString(15.5*cm, ypos, "Estado")
            c.drawString(17.5*cm, ypos, "Fecha Pago")
            ypos -= 0.3*cm
            c.line(2*cm, ypos, ancho - 2*cm, ypos)
            ypos -= 0.4*cm
            c.setFont("Helvetica", 7)
        
        estado = "PAGADO" if campesino['pagado'] else "PENDIENTE"
        fecha_pago = campesino['fecha_pago'] if campesino['fecha_pago'] else "-"
        
        c.drawString(2*cm, ypos, campesino['numero_lote'])
        nombre = campesino['nombre_campesino'][:30]
        c.drawString(3.5*cm, ypos, nombre)
        c.drawString(11*cm, ypos, campesino['barrio'])
        c.drawRightString(15*cm, ypos, f"${campesino['monto']:.2f}")
        c.drawString(15.5*cm, ypos, estado)
        c.drawString(17.5*cm, ypos, fecha_pago[:10] if fecha_pago != "-" else "-")
        
        ypos -= 0.35*cm
    
    # Línea final y totales
    ypos -= 0.2*cm
    c.line(2*cm, ypos, ancho - 2*cm, ypos)
    
    ypos -= 0.5*cm
    c.setFont("Helvetica-Bold", 10)
    c.drawString(11*cm, ypos, "TOTAL RECAUDADO:")
    c.drawRightString(15*cm, ypos, f"${resumen['monto_recaudado']:.2f}")
    
    ypos -= 0.4*cm
    c.drawString(11*cm, ypos, "TOTAL PENDIENTE:")
    c.drawRightString(15*cm, ypos, f"${resumen['monto_pendiente']:.2f}")
    
    c.save()
    
    print(f"✓ Reporte de cuota generado: {ruta_pdf}")
    
    return ruta_pdf


def generar_reporte_todas_cuotas_pdf() -> str:
    """Genera un reporte PDF con todas las cuotas del sistema"""
    from modules.cuotas import obtener_todas_cuotas_con_estado, obtener_estadisticas_generales_cuotas
    
    cuotas = obtener_todas_cuotas_con_estado()
    stats = obtener_estadisticas_generales_cuotas()
    
    # Crear PDF
    reportes_dir = os.path.join('database', 'reportes')
    os.makedirs(reportes_dir, exist_ok=True)
    
    fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f"reporte_general_cuotas_{fecha_str}.pdf"
    ruta_pdf = os.path.join(reportes_dir, nombre_archivo)
    
    c = canvas.Canvas(ruta_pdf, pagesize=letter)
    ancho, alto = letter
    
    # LOGO
    if os.path.exists(LOGO_PATH):
        try:
            c.drawImage(LOGO_PATH, 2*cm, alto - 3*cm, width=2*cm, height=2*cm, mask='auto')
        except:
            pass
    
    # ENCABEZADO
    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(ancho/2, alto - 2*cm, "REPORTE GENERAL DE CUOTAS DE COOPERACIÓN")
    
    c.setFont("Helvetica", 10)
    c.drawCentredString(ancho/2, alto - 2.8*cm, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    
    # RESUMEN GENERAL
    ypos = alto - 4*cm
    c.setFont("Helvetica-Bold", 13)
    c.drawString(2*cm, ypos, "ESTADÍSTICAS GENERALES")
    
    ypos -= 0.8*cm
    c.setFont("Helvetica", 10)
    c.drawString(2*cm, ypos, f"Total de Tipos de Cuotas: {stats['total_tipos_cuotas']}")
    ypos -= 0.5*cm
    c.drawString(2*cm, ypos, f"Total de Cuotas Asignadas: {stats['total_cuotas_asignadas']}")
    ypos -= 0.5*cm
    c.drawString(2*cm, ypos, f"Cuotas Pagadas: {stats['total_pagadas']} - Monto Recaudado: ${stats['monto_recaudado']:.2f}")
    ypos -= 0.5*cm
    c.drawString(2*cm, ypos, f"Cuotas Pendientes: {stats['total_pendientes']} - Monto Pendiente: ${stats['monto_pendiente']:.2f}")
    ypos -= 0.5*cm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2*cm, ypos, f"MONTO TOTAL: ${stats['monto_total']:.2f}")
    
    # TABLA DE CUOTAS
    ypos -= 1.2*cm
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, ypos, "DETALLE POR CUOTA")
    
    ypos -= 0.6*cm
    
    # Encabezados
    c.setFont("Helvetica-Bold", 8)
    c.drawString(2*cm, ypos, "Nombre")
    c.drawString(8*cm, ypos, "Monto Unit.")
    c.drawString(11*cm, ypos, "Asignados")
    c.drawString(13*cm, ypos, "Pagados")
    c.drawString(15*cm, ypos, "Pendientes")
    c.drawString(17*cm, ypos, "Recaudado")
    
    ypos -= 0.3*cm
    c.line(2*cm, ypos, ancho - 2*cm, ypos)
    ypos -= 0.4*cm
    
    # Datos
    c.setFont("Helvetica", 8)
    
    for cuota in cuotas:
        if ypos < 3*cm:
            c.showPage()
            ypos = alto - 2*cm
            c.setFont("Helvetica-Bold", 8)
            c.drawString(2*cm, ypos, "Nombre")
            c.drawString(8*cm, ypos, "Monto Unit.")
            c.drawString(11*cm, ypos, "Asignados")
            c.drawString(13*cm, ypos, "Pagados")
            c.drawString(15*cm, ypos, "Pendientes")
            c.drawString(17*cm, ypos, "Recaudado")
            ypos -= 0.3*cm
            c.line(2*cm, ypos, ancho - 2*cm, ypos)
            ypos -= 0.4*cm
            c.setFont("Helvetica", 8)
        
        c.drawString(2*cm, ypos, cuota['nombre'][:40])
        c.drawRightString(10*cm, ypos, f"${cuota['monto']:.2f}")
        c.drawCentredString(11.5*cm, ypos, str(cuota['total_asignados'] or 0))
        c.drawCentredString(13.5*cm, ypos, str(cuota['total_pagados'] or 0))
        c.drawCentredString(15.5*cm, ypos, str(cuota['total_pendientes'] or 0))
        c.drawRightString(19*cm, ypos, f"${cuota['monto_recaudado'] or 0:.2f}")
        
        ypos -= 0.4*cm
    
    c.save()
    
    print(f"✓ Reporte general de cuotas generado: {ruta_pdf}")
    
    return ruta_pdf

def generar_reporte_cuotas_dia_pdf(fecha: Optional[str] = None) -> str:
    """Genera un reporte PDF de las cuotas cobradas en un día específico"""
    from modules.cuotas import obtener_recibos_cuotas_dia
    
    if not fecha:
        fecha = datetime.now().strftime('%Y-%m-%d')
    
    recibos = obtener_recibos_cuotas_dia(fecha)
    
    if not recibos:
        raise ValueError("No hay recibos de cuotas para el día seleccionado")
    
    # Crear PDF
    reportes_dir = os.path.join('database', 'reportes')
    os.makedirs(reportes_dir, exist_ok=True)
    
    fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f"recaudacion_cuotas_{fecha_str}.pdf"
    ruta_pdf = os.path.join(reportes_dir, nombre_archivo)
    
    c = canvas.Canvas(ruta_pdf, pagesize=letter)
    ancho, alto = letter
    
    # LOGO
    if os.path.exists(LOGO_PATH):
        try:
            c.drawImage(LOGO_PATH, 2*cm, alto - 3*cm, width=2*cm, height=2*cm, mask='auto')
        except:
            pass
    
    # ENCABEZADO
    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(ancho/2, alto - 2*cm, "RECAUDACIÓN DE CUOTAS DEL DÍA")
    
    fecha_obj = datetime.strptime(fecha, '%Y-%m-%d')
    c.setFont("Helvetica", 12)
    c.drawCentredString(ancho/2, alto - 2.8*cm, f"Fecha: {fecha_obj.strftime('%d/%m/%Y')}")
    
    # RESUMEN
    ypos = alto - 4*cm
    total_recaudado = sum(r['monto'] for r in recibos)
    
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, ypos, "RESUMEN DEL DÍA")
    
    ypos -= 0.8*cm
    c.setFont("Helvetica", 11)
    c.drawString(2*cm, ypos, f"Total de Recibos: {len(recibos)}")
    ypos -= 0.5*cm
    c.setFont("Helvetica-Bold", 14)
    c.drawString(2*cm, ypos, f"Total Recaudado: ${total_recaudado:.2f}")
    
    # TABLA
    ypos -= 1.2*cm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2*cm, ypos, "DETALLE DE RECIBOS")
    
    ypos -= 0.6*cm
    
    # Encabezados de tabla
    c.setFont("Helvetica-Bold", 9)
    c.drawString(2*cm, ypos, "Folio")
    c.drawString(3.5*cm, ypos, "Hora")
    c.drawString(5.5*cm, ypos, "Lote")
    c.drawString(7.5*cm, ypos, "Nombre")
    c.drawString(13*cm, ypos, "Barrio")
    c.drawString(15.5*cm, ypos, "Cuota")
    c.drawString(18.5*cm, ypos, "Monto")
    
    ypos -= 0.3*cm
    c.line(2*cm, ypos, ancho - 2*cm, ypos)
    ypos -= 0.4*cm
    
    # Datos
    c.setFont("Helvetica", 8)
    
    for recibo in recibos:
        if ypos < 3*cm:
            c.showPage()
            ypos = alto - 2*cm
            c.setFont("Helvetica-Bold", 9)
            c.drawString(2*cm, ypos, "Folio")
            c.drawString(3.5*cm, ypos, "Hora")
            c.drawString(5.5*cm, ypos, "Lote")
            c.drawString(7.5*cm, ypos, "Nombre")
            c.drawString(13*cm, ypos, "Barrio")
            c.drawString(15.5*cm, ypos, "Cuota")
            c.drawString(18.5*cm, ypos, "Monto")
            ypos -= 0.3*cm
            c.line(2*cm, ypos, ancho - 2*cm, ypos)
            ypos -= 0.4*cm
            c.setFont("Helvetica", 8)
        
        c.drawString(2*cm, ypos, str(recibo['folio']))
        c.drawString(3.5*cm, ypos, recibo['hora'][:5])
        c.drawString(5.5*cm, ypos, recibo['numero_lote'])
        nombre = recibo['nombre_campesino'][:25]
        c.drawString(7.5*cm, ypos, nombre)
        c.drawString(13*cm, ypos, recibo['barrio'][:15])
        cuota = recibo['nombre_cuota'][:20]
        c.drawString(15.5*cm, ypos, cuota)
        c.drawRightString(20*cm, ypos, f"${recibo['monto']:.2f}")
        
        ypos -= 0.35*cm
    
        # Línea final y total
    ypos -= 0.2*cm
    c.line(2*cm, ypos, ancho - 2*cm, ypos)

    ypos -= 0.6*cm
    c.setFont("Helvetica-Bold", 12)
    # ✅ Cambiar posición del texto para que no se encime
    c.drawString(13*cm, ypos, "TOTAL DEL DÍA:")
    c.drawRightString(20*cm, ypos, f"${total_recaudado:.2f}")

    # Pie de página
    ypos -= 1.5*cm
    c.setFont("Helvetica", 8)
    c.drawString(2*cm, ypos, f"Reporte generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}")
    c.drawString(2*cm, ypos - 0.3*cm, "Sistema de Control de Riegos - Módulo de Cuotas de Cooperación")
    
    c.save()
    
    print(f"✓ Reporte de cuotas del día generado: {ruta_pdf}")
    
    return ruta_pdf

def generar_excel_cuotas_dia(fecha: Optional[str] = None) -> str:
    """Genera un Excel con las cuotas cobradas en un día específico"""
    from modules.cuotas import obtener_recibos_cuotas_dia
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    if not fecha:
        fecha = datetime.now().strftime('%Y-%m-%d')
    
    recibos = obtener_recibos_cuotas_dia(fecha)
    
    if not recibos:
        raise ValueError("No hay recibos de cuotas para el día seleccionado")
    
    # Crear Excel
    reportes_dir = os.path.join('database', 'reportes')
    os.makedirs(reportes_dir, exist_ok=True)
    
    fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f"cuotas_dia_{fecha_str}.xlsx"
    ruta_excel = os.path.join(reportes_dir, nombre_archivo)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cuotas del Día"
    
    # ESTILOS
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # TÍTULO
    ws.merge_cells('A1:H1')
    ws['A1'] = "RECAUDACIÓN DE CUOTAS DEL DÍA"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    fecha_obj = datetime.strptime(fecha, '%Y-%m-%d')
    ws.merge_cells('A2:H2')
    ws['A2'] = f"Fecha: {fecha_obj.strftime('%d/%m/%Y')}"
    ws['A2'].font = Font(size=12)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # RESUMEN
    total_recaudado = sum(r['monto'] for r in recibos)
    ws['A4'] = "Total de Recibos:"
    ws['B4'] = len(recibos)
    ws['B4'].font = Font(bold=True)
    
    ws['A5'] = "Total Recaudado:"
    ws['B5'] = total_recaudado
    ws['B5'].font = Font(bold=True)
    ws['B5'].number_format = '"$"#,##0.00'
    
    # ENCABEZADOS
    headers = ['Folio', 'Hora', 'Lote', 'Nombre', 'Barrio', 'Cuota', 'Monto', 'Fecha']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # DATOS
    for idx, recibo in enumerate(recibos, start=8):
        ws.cell(row=idx, column=1, value=recibo['folio'])
        ws.cell(row=idx, column=2, value=recibo['hora'][:5])  # HH:MM
        ws.cell(row=idx, column=3, value=recibo['numero_lote'])
        ws.cell(row=idx, column=4, value=recibo['nombre_campesino'])
        ws.cell(row=idx, column=5, value=recibo['barrio'])
        ws.cell(row=idx, column=6, value=recibo['nombre_cuota'])
        ws.cell(row=idx, column=7, value=recibo['monto'])
        ws.cell(row=idx, column=8, value=recibo['fecha'])
        
        # Formato moneda
        ws.cell(row=idx, column=7).number_format = '"$"#,##0.00'
        
        # Bordes
        for col in range(1, 9):
            ws.cell(row=idx, column=col).border = border
    
    # TOTAL AL FINAL
    ultima_fila = len(recibos) + 8
    ws.merge_cells(f'A{ultima_fila}:F{ultima_fila}')
    ws[f'A{ultima_fila}'] = "TOTAL DEL DÍA:"
    ws[f'A{ultima_fila}'].font = Font(bold=True, size=12)
    ws[f'A{ultima_fila}'].alignment = Alignment(horizontal='right')
    
    ws[f'G{ultima_fila}'] = total_recaudado
    ws[f'G{ultima_fila}'].font = Font(bold=True, size=12)
    ws[f'G{ultima_fila}'].number_format = '"$"#,##0.00'
    ws[f'G{ultima_fila}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # AJUSTAR ANCHOS
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    
    wb.save(ruta_excel)
    
    print(f"✓ Excel de cuotas del día generado: {ruta_excel}")
    
    return ruta_excel
